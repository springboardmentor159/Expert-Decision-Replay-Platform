import io
import uuid
from datetime import datetime, timedelta
import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.main import app

client = TestClient(app)


def get_user_token(
    email: str = "admin_reports@example.com",
    role: str = "Administrator",
    full_name: str = "Report Admin",
    department: str = "Engineering"
):
    emp_id = f"EMP-{uuid.uuid4().hex[:6]}"
    client.post(
        "/users",
        json={
            "email": email,
            "password": "Password123!",
            "full_name": full_name,
            "role": role,
            "employee_id": emp_id,
            "department": department
        }
    )
    resp = client.post(
        "/auth/login",
        json={
            "email": email,
            "password": "Password123!"
        }
    )
    if resp.status_code != 200:
        raise Exception(f"Login failed: {resp.text}")
    return resp.json()["access_token"]


# =============================================================================
# 1. AUTHENTICATION & AUTHORIZATION TESTS (401 & 403)
# =============================================================================

def test_reports_unauthenticated():
    endpoints = [
        "/reports/decisions",
        "/reports/decisions/export/pdf",
        "/reports/decisions/export/excel",
        "/reports/approvals",
        "/reports/approvals/export/pdf",
        "/reports/approvals/export/excel",
        "/reports/teams",
        "/reports/teams/export/pdf",
        "/reports/teams/export/excel",
        "/reports/audit",
        "/reports/audit/export/pdf",
        "/reports/audit/export/excel",
    ]
    for ep in endpoints:
        resp = client.get(ep)
        assert resp.status_code == 401, f"Expected 401 for {ep}, got {resp.status_code}"


def test_reports_role_authorization():
    emp_token = get_user_token("emp_report@example.com", "Employee", "Emp User", "Engineering")
    mgr_eng_token = get_user_token("mgr_eng@example.com", "Manager", "Manager Eng", "Engineering")
    admin_token = get_user_token("admin_rbac@example.com", "Administrator", "Admin User", "Executive")

    emp_hdr = {"Authorization": f"Bearer {emp_token}"}
    mgr_hdr = {"Authorization": f"Bearer {mgr_eng_token}"}
    admin_hdr = {"Authorization": f"Bearer {admin_token}"}

    # 1. Audit reports: Forbidden for Employee and Manager, Allowed for Administrator
    assert client.get("/reports/audit", headers=emp_hdr).status_code == 403
    assert client.get("/reports/audit", headers=mgr_hdr).status_code == 403
    assert client.get("/reports/audit/export/pdf", headers=emp_hdr).status_code == 403
    assert client.get("/reports/audit/export/excel", headers=emp_hdr).status_code == 403
    assert client.get("/reports/audit", headers=admin_hdr).status_code == 200

    # 2. Team reports: Forbidden for Employee, Allowed for Manager (own team) & Admin
    assert client.get("/reports/teams", headers=emp_hdr).status_code == 403
    assert client.get("/reports/teams/export/pdf", headers=emp_hdr).status_code == 403

    # Manager accessing own team -> 200
    mgr_own_resp = client.get("/reports/teams?team=Engineering", headers=mgr_hdr)
    assert mgr_own_resp.status_code == 200

    # Manager attempting to access different team -> 403
    mgr_diff_resp = client.get("/reports/teams?team=Marketing", headers=mgr_hdr)
    assert mgr_diff_resp.status_code == 403

    # Admin accessing any team -> 200
    admin_team_resp = client.get("/reports/teams?team=Marketing", headers=admin_hdr)
    assert admin_team_resp.status_code == 200


# =============================================================================
# 2. VALIDATION ERROR HANDLING (422)
# =============================================================================

def test_reports_validation_errors():
    admin_token = get_user_token("admin_val@example.com", "Administrator", "Admin Val", "IT")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # 1. Invalid date format
    resp = client.get("/reports/decisions?start_date=2026-13-45", headers=headers)
    assert resp.status_code == 422
    assert "start_date" in resp.json()["detail"].lower()

    # 2. Invalid date range (start > end)
    resp = client.get("/reports/approvals?start_date=2026-12-31&end_date=2026-01-01", headers=headers)
    assert resp.status_code == 422
    assert "start_date cannot be after end_date" in resp.json()["detail"]

    # 3. Invalid status
    resp = client.get("/reports/decisions?status=NonExistentStatus", headers=headers)
    assert resp.status_code == 422
    assert "invalid status" in resp.json()["detail"].lower()

    # 4. Invalid sort field
    resp = client.get("/reports/decisions?sort_by=unsupported_column", headers=headers)
    assert resp.status_code == 422
    assert "invalid sort_by" in resp.json()["detail"].lower()

    # 5. Invalid pagination values
    resp = client.get("/reports/decisions?page=0", headers=headers)
    assert resp.status_code == 422

    resp = client.get("/reports/decisions?page_size=0", headers=headers)
    assert resp.status_code == 422

    resp = client.get("/reports/decisions?page_size=200", headers=headers)
    assert resp.status_code == 422


# =============================================================================
# 3. DECISION REPORTS & FILTERING
# =============================================================================

def test_decision_reports_workflow():
    admin_token = get_user_token("admin_dec_rep@example.com", "Administrator", "Dec Admin", "Engineering")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # Create Decisions
    d1 = client.post(
        "/decisions",
        headers=headers,
        json={"title": "Cloud Migration Alpha", "problem_statement": "Migrate on-prem servers", "category": "Infrastructure"}
    ).json()
    d1_id = d1["id"]

    d2 = client.post(
        "/decisions",
        headers=headers,
        json={"title": "Frontend Framework Beta", "problem_statement": "Choose React or Vue", "category": "Frontend"}
    ).json()
    d2_id = d2["id"]

    # Add Alternative to d1
    alt_resp = client.post(
        f"/decisions/{d1_id}/alternatives",
        headers=headers,
        json={
            "name": "AWS EC2",
            "description": "Virtual machines",
            "pros": "Highly scalable and reliable",
            "cons": "Higher running cost",
            "estimated_cost": 1000.0,
            "feasibility_score": 5,
            "risk_level": "Low"
        }
    )
    assert alt_resp.status_code == 201

    # Add Approval to d1
    rev_token = get_user_token("rev_dec_rep@example.com", "Reviewer", "Dec Reviewer", "Engineering")
    rev_user = client.get("/users/me", headers={"Authorization": f"Bearer {rev_token}"}).json()
    client.post(
        "/approvals",
        headers=headers,
        json={"decision_id": d1_id, "reviewer_id": rev_user["id"], "approval_level": 1, "comments": "Initial review"}
    )

    # 1. Fetch Decision Report
    resp = client.get("/reports/decisions", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "summary" in data
    assert data["total"] >= 2
    assert data["summary"]["total_decisions"] >= 2

    # Verify item contents
    item_map = {item["decision_id"]: item for item in data["items"]}
    assert d1_id in item_map
    assert item_map[d1_id]["title"] == "Cloud Migration Alpha"
    assert item_map[d1_id]["category"] == "Infrastructure"
    assert item_map[d1_id]["number_of_alternatives"] == 1
    assert item_map[d1_id]["number_of_approvals"] == 1
    assert item_map[d1_id]["created_by_name"] == "Dec Admin"

    # 2. Filter by Category
    cat_resp = client.get("/reports/decisions?category=Infrastructure", headers=headers)
    assert cat_resp.status_code == 200
    cat_data = cat_resp.json()
    assert all(it["category"] == "Infrastructure" for it in cat_data["items"])

    # 3. Filter by Status
    stat_resp = client.get("/reports/decisions?status=Draft", headers=headers)
    assert stat_resp.status_code == 200
    assert all(it["status"] == "Draft" for it in stat_resp.json()["items"])

    # 4. Controlled Sorting
    sort_resp = client.get("/reports/decisions?sort_by=title&sort_order=asc", headers=headers)
    assert sort_resp.status_code == 200
    titles = [it["title"] for it in sort_resp.json()["items"]]
    assert titles == sorted(titles)

    # 5. Pagination
    page_resp = client.get("/reports/decisions?page=1&page_size=1", headers=headers)
    assert page_resp.status_code == 200
    page_data = page_resp.json()
    assert len(page_data["items"]) == 1
    assert page_data["page"] == 1
    assert page_data["page_size"] == 1
    assert page_data["total_pages"] >= 2


# =============================================================================
# 4. APPROVAL REPORTS & TURNAROUND METRICS
# =============================================================================

def test_approval_reports_workflow():
    admin_token = get_user_token("admin_appr_rep@example.com", "Administrator", "Appr Admin", "Operations")
    headers = {"Authorization": f"Bearer {admin_token}"}

    rev_token = get_user_token("rev_appr_rep@example.com", "Reviewer", "Appr Reviewer", "Operations")
    rev_hdr = {"Authorization": f"Bearer {rev_token}"}
    rev_user = client.get("/users/me", headers=rev_hdr).json()

    # Create Decision
    d = client.post(
        "/decisions",
        headers=headers,
        json={"title": "Security Protocol Gamma", "problem_statement": "Zero Trust upgrade", "category": "Security"}
    ).json()

    # Create 2 Approvals
    app1 = client.post(
        "/approvals",
        headers=headers,
        json={"decision_id": d["id"], "reviewer_id": rev_user["id"], "approval_level": 1, "comments": "Review 1"}
    ).json()

    app2 = client.post(
        "/approvals",
        headers=headers,
        json={"decision_id": d["id"], "reviewer_id": rev_user["id"], "approval_level": 2, "comments": "Review 2"}
    ).json()

    # Approve app1
    client.post(
        f"/approvals/{app1['id']}/approve",
        headers=rev_hdr,
        json={"comments": "LGTM"}
    )

    # Reject app2
    client.post(
        f"/approvals/{app2['id']}/reject",
        headers=rev_hdr,
        json={"comments": "Needs revisions"}
    )

    # Fetch Approval Report
    resp = client.get("/reports/approvals", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "summary" in data
    assert data["summary"]["total_approvals"] >= 2
    assert data["summary"]["approved_approvals"] >= 1
    assert data["summary"]["rejected_approvals"] >= 1
    assert data["summary"]["approval_completion_rate"] > 0.0

    # Verify turnaround calculation
    item_map = {it["approval_id"]: it for it in data["items"]}
    assert app1["id"] in item_map
    assert item_map[app1["id"]]["approval_status"] == "Approved"
    assert item_map[app1["id"]]["turnaround_time_hours"] is not None
    assert item_map[app1["id"]]["turnaround_time_hours"] >= 0.0

    # Filter by Status
    app_filt_resp = client.get("/reports/approvals?status=Approved", headers=headers)
    assert app_filt_resp.status_code == 200
    assert all(it["approval_status"] == "Approved" for it in app_filt_resp.json()["items"])


# =============================================================================
# 5. TEAM REPORTS & RBAC SCOPING
# =============================================================================

def test_team_reports_workflow():
    admin_token = get_user_token("admin_team_rep@example.com", "Administrator", "Team Admin", "HQ")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # Create users in different departments
    get_user_token("eng_user1@example.com", "Employee", "Eng 1", "Engineering")
    get_user_token("eng_user2@example.com", "Employee", "Eng 2", "Engineering")
    get_user_token("sales_user@example.com", "Employee", "Sales 1", "Sales")

    # Fetch Team Report as Admin
    resp = client.get("/reports/teams", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "summary" in data
    assert data["summary"]["total_teams"] >= 2
    assert data["summary"]["total_members"] >= 3

    team_map = {t["team_name"]: t for t in data["items"]}
    assert "Engineering" in team_map
    assert team_map["Engineering"]["number_of_members"] >= 2

    # Test sorting
    sort_team = client.get("/reports/teams?sort_by=number_of_members&sort_order=desc", headers=headers)
    assert sort_team.status_code == 200
    counts = [t["number_of_members"] for t in sort_team.json()["items"]]
    assert counts == sorted(counts, reverse=True)


# =============================================================================
# 6. AUDIT REPORTS & RBAC
# =============================================================================

def test_audit_reports_workflow():
    admin_token = get_user_token("admin_audit_rep@example.com", "Administrator", "Audit Admin", "Executive")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # Perform action to generate audit log
    client.post(
        "/decisions",
        headers=headers,
        json={"title": "Audit Test Decision", "problem_statement": "Testing audit", "category": "General"}
    )

    # Fetch Audit Report
    resp = client.get("/reports/audit", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert "items" in data
    assert "summary" in data
    assert data["total"] >= 1
    assert "CREATE" in data["summary"]["action_breakdown"]

    # Filter by action
    act_resp = client.get("/reports/audit?action=CREATE", headers=headers)
    assert act_resp.status_code == 200
    assert all(it["action"] == "CREATE" for it in act_resp.json()["items"])

    # Filter by entity_type
    ent_resp = client.get("/reports/audit?entity_type=Decision", headers=headers)
    assert ent_resp.status_code == 200
    assert all(it["entity_type"].lower() == "decision" for it in ent_resp.json()["items"])


# =============================================================================
# 7. PDF EXPORT VALIDATION
# =============================================================================

def test_pdf_exports():
    admin_token = get_user_token("admin_pdf_exp@example.com", "Administrator", "PDF Admin", "Design")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # 1. Decisions PDF Export
    resp_dec = client.get("/reports/decisions/export/pdf", headers=headers)
    assert resp_dec.status_code == 200
    assert resp_dec.headers["content-type"] == "application/pdf"
    assert 'attachment; filename="decisions_report.pdf"' in resp_dec.headers["content-disposition"]
    assert resp_dec.content.startswith(b"%PDF-")

    # 2. Approvals PDF Export
    resp_app = client.get("/reports/approvals/export/pdf", headers=headers)
    assert resp_app.status_code == 200
    assert resp_app.headers["content-type"] == "application/pdf"
    assert 'attachment; filename="approvals_report.pdf"' in resp_app.headers["content-disposition"]
    assert resp_app.content.startswith(b"%PDF-")

    # 3. Teams PDF Export
    resp_team = client.get("/reports/teams/export/pdf", headers=headers)
    assert resp_team.status_code == 200
    assert resp_team.headers["content-type"] == "application/pdf"
    assert 'attachment; filename="teams_report.pdf"' in resp_team.headers["content-disposition"]
    assert resp_team.content.startswith(b"%PDF-")

    # 4. Audit PDF Export
    resp_audit = client.get("/reports/audit/export/pdf", headers=headers)
    assert resp_audit.status_code == 200
    assert resp_audit.headers["content-type"] == "application/pdf"
    assert 'attachment; filename="audit_report.pdf"' in resp_audit.headers["content-disposition"]
    assert resp_audit.content.startswith(b"%PDF-")


# =============================================================================
# 8. EXCEL EXPORT VALIDATION
# =============================================================================

def test_excel_exports():
    admin_token = get_user_token("admin_xlsx_exp@example.com", "Administrator", "XLSX Admin", "Analytics")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # 1. Decisions Excel Export
    resp_dec = client.get("/reports/decisions/export/excel", headers=headers)
    assert resp_dec.status_code == 200
    assert "openxmlformats-officedocument.spreadsheetml.sheet" in resp_dec.headers["content-type"]
    assert 'attachment; filename="decisions_report.xlsx"' in resp_dec.headers["content-disposition"]
    
    wb_dec = openpyxl.load_workbook(io.BytesIO(resp_dec.content))
    assert "Summary & Filters" in wb_dec.sheetnames
    assert "Decisions" in wb_dec.sheetnames

    # 2. Approvals Excel Export
    resp_app = client.get("/reports/approvals/export/excel", headers=headers)
    assert resp_app.status_code == 200
    wb_app = openpyxl.load_workbook(io.BytesIO(resp_app.content))
    assert "Summary & Filters" in wb_app.sheetnames
    assert "Approvals" in wb_app.sheetnames

    # 3. Teams Excel Export
    resp_team = client.get("/reports/teams/export/excel", headers=headers)
    assert resp_team.status_code == 200
    wb_team = openpyxl.load_workbook(io.BytesIO(resp_team.content))
    assert "Summary & Filters" in wb_team.sheetnames
    assert "Teams" in wb_team.sheetnames

    # 4. Audit Excel Export
    resp_audit = client.get("/reports/audit/export/excel", headers=headers)
    assert resp_audit.status_code == 200
    wb_audit = openpyxl.load_workbook(io.BytesIO(resp_audit.content))
    assert "Summary & Filters" in wb_audit.sheetnames
    assert "Audit Trail" in wb_audit.sheetnames


# =============================================================================
# 9. EMPTY MATCHES HANDLING
# =============================================================================

def test_empty_matches_handling():
    admin_token = get_user_token("admin_empty@example.com", "Administrator", "Empty Admin", "Test")
    headers = {"Authorization": f"Bearer {admin_token}"}

    # Query with non-matching filter
    resp = client.get("/reports/decisions?category=NonExistentCategoryXYZ", headers=headers)
    assert resp.status_code == 200
    data = resp.json()
    assert data["items"] == []
    assert data["total"] == 0
    assert data["summary"]["total_decisions"] == 0

    # Export empty PDF
    pdf_resp = client.get("/reports/decisions/export/pdf?category=NonExistentCategoryXYZ", headers=headers)
    assert pdf_resp.status_code == 200
    assert pdf_resp.content.startswith(b"%PDF-")

    # Export empty Excel
    xlsx_resp = client.get("/reports/decisions/export/excel?category=NonExistentCategoryXYZ", headers=headers)
    assert xlsx_resp.status_code == 200
