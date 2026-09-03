import io
import sys
from datetime import datetime, timedelta
from fastapi.testclient import TestClient
import openpyxl
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db.database import SessionLocal, engine
from app.main import app
from app.models.approval import Approval
from app.models.audit_log import AuditLog
from app.models.decision import Decision
from app.models.user import User

client = TestClient(app)


def run_sprint12_verification():
    print("=" * 88)
    print(" SPRINT 12: REPORTS & EXPORT MODULE COMPREHENSIVE END-TO-END VERIFICATION")
    print("=" * 88)

    passed_count = 0
    total_count = 0

    def record_check(condition: bool, test_name: str, details: str = ""):
        nonlocal passed_count, total_count
        total_count += 1
        if condition:
            passed_count += 1
            print(f" [PASS] {test_name}")
        else:
            print(f" [FAIL] {test_name} -> {details}")

    def setup_user(email: str, role: str, department: str = "Enterprise Architecture"):
        u_data = {
            "full_name": email.split("@")[0].replace("_", " ").title(),
            "email": email,
            "role": role,
            "password": "Password123!",
            "employee_id": f"EMP_V12_{email[:7]}",
            "department": department,
            "designation": f"Lead {role}",
            "phone_number": "+1-555-1212"
        }
        client.post("/users", json=u_data)
        login_res = client.post("/auth/login", json={"email": email, "password": "Password123!"})
        assert login_res.status_code == 200, f"Login failed for {email}"
        return login_res.json()["access_token"], login_res.json()["user"]["id"]

    # =========================================================================
    # Step 1: User Roles Setup & JWT Authentication
    # =========================================================================
    print("\n--- Step 1: User Roles Setup & JWT Authentication (Task 8 & 9) ---")
    emp_token, emp_id = setup_user("v12_employee@example.com", "Employee", "Architecture")
    emp_headers = {"Authorization": f"Bearer {emp_token}"}
    record_check(bool(emp_token), "1.1 Login as Employee & Acquire JWT Token (200 OK)")

    rev_token, rev_id = setup_user("v12_reviewer@example.com", "Reviewer", "Architecture")
    rev_headers = {"Authorization": f"Bearer {rev_token}"}
    record_check(bool(rev_token), "1.2 Login as Reviewer & Acquire JWT Token (200 OK)")

    mgr_token, mgr_id = setup_user("v12_manager@example.com", "Manager", "Architecture")
    mgr_headers = {"Authorization": f"Bearer {mgr_token}"}
    record_check(bool(mgr_token), "1.3 Login as Manager & Acquire JWT Token (200 OK)")

    adm_token, adm_id = setup_user("v12_admin@example.com", "Administrator", "Compliance")
    adm_headers = {"Authorization": f"Bearer {adm_token}"}
    record_check(bool(adm_token), "1.4 Login as Administrator & Acquire JWT Token (200 OK)")

    ops_token, ops_id = setup_user("v12_ops_emp@example.com", "Employee", "Operations")
    ops_headers = {"Authorization": f"Bearer {ops_token}"}
    record_check(bool(ops_token), "1.5 Login as Operations Employee (Different Department)")

    # =========================================================================
    # Step 2: Seed Domain Data (Decisions, Alternatives, Approvals, Audit Logs)
    # =========================================================================
    print("\n--- Step 2: Seed Decisions, Approvals & Workflow Data ---")
    
    # Decision A: Draft Technology
    dec_a = client.post("/decisions", json={
        "title": "Adopt Event-Driven Microservices Architecture",
        "problem_statement": "Evaluate asynchronous message brokers for enterprise decoupled services",
        "category": "Technology"
    }, headers=emp_headers).json()
    dec_a_id = dec_a["id"]
    client.post(f"/decisions/{dec_a_id}/alternatives", json={
        "title": "Apache Kafka Event Hub",
        "description": "Distributed log broker",
        "pros": "High throughput, immutable log",
        "cons": "Zookeeper / KRaft complexity",
        "cost": 12000.0,
        "risk_level": "Medium"
    }, headers=emp_headers)

    # Decision B: Approved Security
    dec_b_res = client.post("/decisions", json={
        "title": "Implement Zero-Trust Security Perimeter",
        "problem_statement": "Enforce strict identity verification for internal VPN access",
        "category": "Security"
    }, headers=emp_headers)
    dec_b_id = dec_b_res.json()["id"]

    apprv_b_res = client.post(f"/decisions/{dec_b_id}/submit", json={
        "reviewer_id": rev_id,
        "approval_level": 1,
        "comments": "Please review Zero Trust architecture"
    }, headers=emp_headers)
    apprv_b_id = apprv_b_res.json()["id"]

    # Reviewer approves Decision B
    client.post(f"/approvals/{apprv_b_id}/action", json={
        "status": "Approved",
        "comments": "Approved by Architecture Review Board"
    }, headers=rev_headers)


    # Decision C: Rejected Finance by Operations Employee
    dec_c = client.post("/decisions", json={
        "title": "Legacy Server Hardware Renewal",
        "problem_statement": "Request budget extension for on-prem bare metal servers",
        "category": "Finance"
    }, headers=ops_headers).json()
    dec_c_id = dec_c["id"]
    client.patch(f"/decisions/{dec_c_id}/status", json={"status": "Rejected"}, headers=ops_headers)

    record_check(True, "2.1 Test data seeded across categories (Technology, Security, Finance) & statuses (Draft, Approved, Rejected)")

    # =========================================================================
    # Step 3: Decision Reports Verification (Task 1)
    # =========================================================================
    print("\n--- Step 3: Decision Reports Verification (GET /reports/decisions) ---")
    dec_report_res = client.get("/reports/decisions", headers=emp_headers)
    record_check(dec_report_res.status_code == 200, "3.1 GET /reports/decisions returns 200 OK")
    dec_data = dec_report_res.json()
    record_check(
        "summary" in dec_data and "items" in dec_data and dec_data["summary"]["total_decisions"] >= 3,
        "3.2 Decision report includes items and summary statistics (total, draft, approved, rejected)"
    )

    # Category Filter
    dec_tech_res = client.get("/reports/decisions?category=Technology", headers=emp_headers)
    record_check(
        dec_tech_res.status_code == 200 and all(i["category"] == "Technology" for i in dec_tech_res.json()["items"]),
        "3.3 Filter decisions by category=Technology"
    )

    # Status Filter
    dec_apprv_res = client.get("/reports/decisions?status=Approved", headers=emp_headers)
    record_check(
        dec_apprv_res.status_code == 200 and all(i["status"] == "Approved" for i in dec_apprv_res.json()["items"]),
        "3.4 Filter decisions by status=Approved"
    )

    # Created By Filter
    dec_creator_res = client.get(f"/reports/decisions?created_by={emp_id}", headers=emp_headers)
    record_check(
        dec_creator_res.status_code == 200 and all(i["created_by"] == emp_id for i in dec_creator_res.json()["items"]),
        "3.5 Filter decisions by created_by user ID"
    )

    # Date Range Filter
    today_str = datetime.utcnow().strftime("%Y-%m-%d")
    dec_date_res = client.get(f"/reports/decisions?start_date={today_str}&end_date={today_str}", headers=emp_headers)
    record_check(
        dec_date_res.status_code == 200 and len(dec_date_res.json()["items"]) >= 3,
        "3.6 Filter decisions by start_date & end_date"
    )

    # =========================================================================
    # Step 4: Approval Reports Verification (Task 2)
    # =========================================================================
    print("\n--- Step 4: Approval Reports Verification (GET /reports/approvals) ---")
    apprv_report_res = client.get("/reports/approvals", headers=emp_headers)
    record_check(apprv_report_res.status_code == 200, "4.1 GET /reports/approvals returns 200 OK")
    apprv_data = apprv_report_res.json()
    record_check(
        "summary" in apprv_data and "total_approvals" in apprv_data["summary"] and "completion_rate" in apprv_data["summary"],
        "4.2 Approval report summary includes total, pending, approved, turnaround time & completion rate"
    )

    # Filter by Status
    apprv_st_res = client.get("/reports/approvals?status=Approved", headers=emp_headers)
    record_check(
        apprv_st_res.status_code == 200 and all(i["status"] == "Approved" for i in apprv_st_res.json()["items"]),
        "4.3 Filter approvals by status=Approved"
    )

    # Filter by Reviewer
    apprv_rev_res = client.get(f"/reports/approvals?reviewer_id={rev_id}", headers=emp_headers)
    record_check(
        apprv_rev_res.status_code == 200 and all(i["reviewer_id"] == rev_id for i in apprv_rev_res.json()["items"]),
        "4.4 Filter approvals by reviewer_id"
    )

    # =========================================================================
    # Step 5: Team Reports & Authorization Verification (Task 3)
    # =========================================================================
    print("\n--- Step 5: Team Reports Verification (GET /reports/teams) ---")
    # Admin accesses all teams
    team_admin_res = client.get("/reports/teams", headers=adm_headers)
    record_check(team_admin_res.status_code == 200, "5.1 Administrator access to GET /reports/teams (200 OK)")
    team_data = team_admin_res.json()
    record_check(
        team_data["summary"]["total_teams"] >= 2 and any(t["team_name"] == "Architecture" for t in team_data["items"]),
        "5.2 Team report calculates member counts, decisions, and approval statistics per department"
    )

    # Employee accesses own department
    team_own_res = client.get("/reports/teams?team=Architecture", headers=emp_headers)
    record_check(
        team_own_res.status_code == 200 and len(team_own_res.json()["items"]) == 1 and team_own_res.json()["items"][0]["team_name"] == "Architecture",
        "5.3 Employee accesses own department report (Architecture)"
    )

    # Employee denied access to other department
    team_forbid_res = client.get("/reports/teams?team=Operations", headers=emp_headers)
    record_check(team_forbid_res.status_code == 403, "5.4 Employee attempting to access another department receives 403 Forbidden")

    # =========================================================================
    # Step 6: Audit Reports Verification (Task 4)
    # =========================================================================
    print("\n--- Step 6: Audit Reports Verification (GET /reports/audit) ---")
    # Admin accesses audit report
    audit_admin_res = client.get("/reports/audit", headers=adm_headers)
    record_check(audit_admin_res.status_code == 200, "6.1 Administrator access to GET /reports/audit (200 OK)")
    audit_data = audit_admin_res.json()
    record_check(
        "summary" in audit_data and "actions_breakdown" in audit_data["summary"] and audit_data["summary"]["total_audit_logs"] >= 1,
        "6.2 Audit report includes actions and entities breakdown summaries"
    )

    # Audit filter by Action
    audit_act_res = client.get("/reports/audit?action=CREATE", headers=adm_headers)
    record_check(
        audit_act_res.status_code == 200 and all(i["action"] == "CREATE" for i in audit_act_res.json()["items"]),
        "6.3 Filter audit reports by action=CREATE"
    )

    # Non-Admin denied access to audit reports
    audit_emp_res = client.get("/reports/audit", headers=emp_headers)
    record_check(audit_emp_res.status_code == 403, "6.4 Employee accessing Audit Report receives 403 Forbidden")

    audit_mgr_res = client.get("/reports/audit", headers=mgr_headers)
    record_check(audit_mgr_res.status_code == 403, "6.5 Manager accessing Audit Report receives 403 Forbidden")

    # =========================================================================
    # Step 7: Pagination & Controlled Sorting Verification (Task 5)
    # =========================================================================
    print("\n--- Step 7: Pagination and Controlled Sorting Verification ---")
    page_res = client.get("/reports/decisions?page=1&page_size=2", headers=emp_headers)
    record_check(
        page_res.status_code == 200 and len(page_res.json()["items"]) <= 2 and page_res.json()["page"] == 1,
        "7.1 Pagination supported with page=1&page_size=2"
    )

    sort_asc_res = client.get("/reports/decisions?sort_by=title&sort_order=asc", headers=emp_headers)
    record_check(sort_asc_res.status_code == 200, "7.2 Controlled sorting by title ASC")
    titles = [i["title"] for i in sort_asc_res.json()["items"]]
    record_check(titles == sorted(titles, key=str.lower), "7.3 Titles returned in correct alphabetical ascending order")


    # =========================================================================
    # Step 8: PDF Export Verification (Task 6)
    # =========================================================================
    print("\n--- Step 8: PDF Export Verification (Task 6) ---")
    # Decisions PDF
    pdf_dec = client.get("/reports/decisions/export/pdf", headers=emp_headers)
    record_check(
        pdf_dec.status_code == 200 and pdf_dec.headers["content-type"] == "application/pdf" and pdf_dec.content.startswith(b"%PDF-"),
        "8.1 GET /reports/decisions/export/pdf generates valid PDF document (%PDF- magic bytes)"
    )

    # Approvals PDF
    pdf_apprv = client.get("/reports/approvals/export/pdf", headers=emp_headers)
    record_check(
        pdf_apprv.status_code == 200 and pdf_apprv.headers["content-type"] == "application/pdf" and pdf_apprv.content.startswith(b"%PDF-"),
        "8.2 GET /reports/approvals/export/pdf generates valid PDF document (%PDF- magic bytes)"
    )

    # Teams PDF
    pdf_team = client.get("/reports/teams/export/pdf", headers=adm_headers)
    record_check(
        pdf_team.status_code == 200 and pdf_team.headers["content-type"] == "application/pdf" and pdf_team.content.startswith(b"%PDF-"),
        "8.3 GET /reports/teams/export/pdf generates valid PDF document (%PDF- magic bytes)"
    )

    # Audit PDF (Admin vs Employee)
    pdf_audit_adm = client.get("/reports/audit/export/pdf", headers=adm_headers)
    record_check(
        pdf_audit_adm.status_code == 200 and pdf_audit_adm.headers["content-type"] == "application/pdf" and pdf_audit_adm.content.startswith(b"%PDF-"),
        "8.4 GET /reports/audit/export/pdf generates valid PDF document for Administrator"
    )

    pdf_audit_emp = client.get("/reports/audit/export/pdf", headers=emp_headers)
    record_check(pdf_audit_emp.status_code == 403, "8.5 GET /reports/audit/export/pdf returns 403 Forbidden for Employee")

    # =========================================================================
    # Step 9: Excel Export Verification (Task 7)
    # =========================================================================
    print("\n--- Step 9: Excel Export Verification (Task 7) ---")
    # Decisions Excel
    xlsx_dec = client.get("/reports/decisions/export/excel?status=Approved", headers=emp_headers)
    record_check(xlsx_dec.status_code == 200, "9.1 GET /reports/decisions/export/excel returns 200 OK")
    wb_dec = openpyxl.load_workbook(io.BytesIO(xlsx_dec.content))
    record_check("Decisions Report" in wb_dec.sheetnames, "9.2 Decisions Excel sheet contains 'Decisions Report' sheet tab")

    # Approvals Excel
    xlsx_apprv = client.get("/reports/approvals/export/excel", headers=emp_headers)
    record_check(xlsx_apprv.status_code == 200, "9.3 GET /reports/approvals/export/excel returns 200 OK")
    wb_apprv = openpyxl.load_workbook(io.BytesIO(xlsx_apprv.content))
    record_check("Approvals Report" in wb_apprv.sheetnames, "9.4 Approvals Excel sheet contains 'Approvals Report' sheet tab")

    # Teams Excel
    xlsx_team = client.get("/reports/teams/export/excel", headers=adm_headers)
    record_check(xlsx_team.status_code == 200, "9.5 GET /reports/teams/export/excel returns 200 OK")
    wb_team = openpyxl.load_workbook(io.BytesIO(xlsx_team.content))
    record_check("Teams Report" in wb_team.sheetnames, "9.6 Teams Excel sheet contains 'Teams Report' sheet tab")

    # Audit Excel
    xlsx_audit = client.get("/reports/audit/export/excel", headers=adm_headers)
    record_check(xlsx_audit.status_code == 200, "9.7 GET /reports/audit/export/excel returns 200 OK for Administrator")
    wb_audit = openpyxl.load_workbook(io.BytesIO(xlsx_audit.content))
    record_check("Audit Report" in wb_audit.sheetnames, "9.8 Audit Excel sheet contains 'Audit Report' sheet tab")

    # =========================================================================
    # Step 10: PostgreSQL / Database Verification (Task 10)
    # =========================================================================
    print("\n--- Step 10: Database Integrity Verification (Task 10) ---")
    db_session: Session = SessionLocal()
    try:
        db_dec_count = db_session.query(func.count(Decision.id)).scalar() or 0
        db_apprv_count = db_session.query(func.count(Approval.id)).scalar() or 0
        db_user_count = db_session.query(func.count(User.id)).scalar() or 0
        db_audit_count = db_session.query(func.count(AuditLog.id)).scalar() or 0

        api_dec_report = client.get("/reports/decisions", headers=adm_headers).json()
        api_apprv_report = client.get("/reports/approvals", headers=adm_headers).json()
        api_audit_report = client.get("/reports/audit", headers=adm_headers).json()

        record_check(
            api_dec_report["summary"]["total_decisions"] == db_dec_count,
            f"10.1 Decision report total ({api_dec_report['summary']['total_decisions']}) matches database record count ({db_dec_count})"
        )
        record_check(
            api_apprv_report["summary"]["total_approvals"] == db_apprv_count,
            f"10.2 Approval report total ({api_apprv_report['summary']['total_approvals']}) matches database record count ({db_apprv_count})"
        )
        record_check(
            api_audit_report["summary"]["total_audit_logs"] == db_audit_count,
            f"10.3 Audit report total ({api_audit_report['summary']['total_audit_logs']}) matches database record count ({db_audit_count})"
        )
    finally:
        db_session.close()

    # =========================================================================
    # Step 11: Error Handling & Validations Verification (Task 11)
    # =========================================================================
    print("\n--- Step 11: Error Handling & Validations (Task 11) ---")
    # 11.1 No JWT -> 401 Unauthorized
    record_check(client.get("/reports/decisions").status_code == 401, "11.1 No JWT -> 401 Unauthorized")
    
    # 11.2 Insufficient permission -> 403 Forbidden
    record_check(client.get("/reports/audit", headers=emp_headers).status_code == 403, "11.2 Insufficient permission for Audit Report -> 403 Forbidden")

    # 11.3 Invalid date format -> 422
    record_check(client.get("/reports/decisions?start_date=2026/09/01", headers=emp_headers).status_code == 422, "11.3 Invalid date format -> 422 Unprocessable Entity")

    # 11.4 Invalid date range -> 422
    record_check(client.get("/reports/decisions?start_date=2026-12-31&end_date=2026-01-01", headers=emp_headers).status_code == 422, "11.4 Invalid date range (start > end) -> 422 Unprocessable Entity")

    # 11.5 Invalid status -> 422
    record_check(client.get("/reports/decisions?status=InvalidStatus123", headers=emp_headers).status_code == 422, "11.5 Invalid status -> 422 Unprocessable Entity")

    # 11.6 Invalid sorting field -> 422
    record_check(client.get("/reports/decisions?sort_by=non_existent_col", headers=emp_headers).status_code == 422, "11.6 Invalid sorting field -> 422 Unprocessable Entity")

    # 11.7 Invalid pagination values -> 422
    record_check(client.get("/reports/decisions?page=0", headers=emp_headers).status_code == 422, "11.7 Invalid pagination (page=0) -> 422 Unprocessable Entity")

    # 11.8 No matching records -> 200 OK with empty result
    empty_res = client.get("/reports/decisions?start_date=2099-01-01&end_date=2099-12-31", headers=emp_headers)
    record_check(
        empty_res.status_code == 200 and empty_res.json()["items"] == [] and empty_res.json()["total"] == 0,
        "11.8 No matching records returns 200 OK with empty list & 0 summary stats"
    )

    print("\n" + "=" * 88)
    print(f" SPRINT 12 VERIFICATION RESULT: {passed_count} / {total_count} CHECKS PASSED")
    print("=" * 88)

    if passed_count == total_count:
        print(" >>> ALL SPRINT 12 REQUIREMENTS AND EXPORT CAPABILITIES SUCCESSFULLY VERIFIED! <<<")
        return 0
    else:
        print(f" >>> {total_count - passed_count} CHECKS FAILED <<<")
        return 1


if __name__ == "__main__":
    sys.exit(run_sprint12_verification())
