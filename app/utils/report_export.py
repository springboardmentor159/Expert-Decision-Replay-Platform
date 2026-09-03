"""
Sprint 12: Reports & Export - shared PDF/Excel building helpers.

Kept intentionally generic (title + filters + summary + columns + rows)
so every report router (decisions, approvals, teams, audit) can reuse
the exact same PDF/Excel layout instead of hand-rolling one per report.
"""
from datetime import datetime
from io import BytesIO
from typing import Any, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _format_filters(filters: dict[str, Any]) -> str:
    applied = {k: v for k, v in filters.items() if v is not None and v != ""}
    if not applied:
        return "None"
    return ", ".join(f"{k}={v}" for k, v in applied.items())


def build_pdf_report(
    *,
    report_title: str,
    filters: dict[str, Any],
    columns: list[str],
    rows: list[list[Any]],
    summary: Optional[dict[str, Any]] = None,
    generated_at: Optional[datetime] = None,
) -> BytesIO:
    """
    Builds a professionally structured, readable PDF containing:
    report title, generated date, applied filters, summary information,
    and the report data as a table.
    """
    generated_at = generated_at or datetime.utcnow()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        title=report_title,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Title"], fontSize=18, spaceAfter=6
    )
    meta_style = ParagraphStyle(
        "ReportMeta", parent=styles["Normal"], fontSize=9, textColor=colors.grey
    )
    section_style = ParagraphStyle(
        "SectionHeading", parent=styles["Heading2"], fontSize=12, spaceBefore=10, spaceAfter=6
    )

    elements: list = []

    elements.append(Paragraph(report_title, title_style))
    elements.append(
        Paragraph(
            f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')}", meta_style
        )
    )
    elements.append(
        Paragraph(f"Applied filters: {_format_filters(filters)}", meta_style)
    )
    elements.append(Spacer(1, 10))

    if summary:
        elements.append(Paragraph("Summary", section_style))
        summary_rows = [[str(k).replace("_", " ").title(), str(v)] for k, v in summary.items()]
        summary_table = Table(summary_rows, colWidths=[6 * cm, 6 * cm])
        summary_table.setStyle(
            TableStyle(
                [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ("BACKGROUND", (0, 0), (0, -1), colors.whitesmoke),
                    ("FONTSIZE", (0, 0), (-1, -1), 8),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )
        elements.append(summary_table)
        elements.append(Spacer(1, 12))

    elements.append(Paragraph(f"Report Data ({len(rows)} records)", section_style))

    if rows:
        wrap_style = ParagraphStyle("Cell", parent=styles["Normal"], fontSize=7, leading=9)
        header = [Paragraph(f"<b>{c}</b>", wrap_style) for c in columns]
        body = [
            [Paragraph("" if cell is None else str(cell), wrap_style) for cell in row]
            for row in rows
        ]
        table_data = [header] + body

        col_width = (landscape(A4)[0] - 3 * cm) / max(len(columns), 1)
        data_table = Table(table_data, colWidths=[col_width] * len(columns), repeatRows=1)
        data_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.lightgrey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5FA")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("FONTSIZE", (0, 0), (-1, -1), 7),
                ]
            )
        )
        elements.append(data_table)
    else:
        elements.append(Paragraph("No matching records found for the applied filters.", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer


def build_excel_report(
    *,
    report_title: str,
    filters: dict[str, Any],
    columns: list[str],
    rows: list[list[Any]],
    summary: Optional[dict[str, Any]] = None,
    generated_at: Optional[datetime] = None,
) -> BytesIO:
    """
    Builds an Excel workbook with properly structured columns that
    respects whatever filters were applied to the underlying report.
    """
    generated_at = generated_at or datetime.utcnow()

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    title_font = Font(bold=True, size=14)
    meta_font = Font(italic=True, size=9, color="666666")

    row_cursor = 1

    ws.cell(row=row_cursor, column=1, value=report_title).font = title_font
    row_cursor += 1

    ws.cell(
        row=row_cursor, column=1,
        value=f"Generated: {generated_at.strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ).font = meta_font
    row_cursor += 1

    ws.cell(
        row=row_cursor, column=1,
        value=f"Applied filters: {_format_filters(filters)}"
    ).font = meta_font
    row_cursor += 2

    if summary:
        ws.cell(row=row_cursor, column=1, value="Summary").font = Font(bold=True, size=11)
        row_cursor += 1
        for key, value in summary.items():
            ws.cell(row=row_cursor, column=1, value=str(key).replace("_", " ").title())
            ws.cell(row=row_cursor, column=2, value=value)
            row_cursor += 1
        row_cursor += 1

    header_row = row_cursor
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_cursor += 1

    for row in rows:
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_cursor, column=col_idx, value=value)
        row_cursor += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            if col_idx - 1 < len(row) and row[col_idx - 1] is not None:
                max_len = max(max_len, len(str(row[col_idx - 1])))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
