from datetime import datetime
from io import BytesIO
from typing import Optional, List
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_decisions_pdf(report, category, status, start_date, end_date):
    """Generate PDF for decision reports"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph("Decision Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Generated date and filters
    generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"<b>Generated Date:</b> {generated_date}", styles['Normal']))
    
    filter_text = "Filters: "
    if category:
        filter_text += f"Category: {category}, "
    if status:
        filter_text += f"Status: {status}, "
    if start_date:
        filter_text += f"Start Date: {start_date.strftime('%Y-%m-%d')}, "
    if end_date:
        filter_text += f"End Date: {end_date.strftime('%Y-%m-%d')}, "
    
    if filter_text == "Filters: ":
        filter_text += "None"
    else:
        filter_text = filter_text.rstrip(", ")
    
    elements.append(Paragraph(f"<b>{filter_text}</b>", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Summary
    summary = report.summary
    summary_text = f"""
    <b>Summary Statistics:</b><br/>
    Total Decisions: {summary.total_decisions} | 
    Draft: {summary.draft_decisions} | 
    Under Review: {summary.decisions_under_review} | 
    Approved: {summary.approved_decisions} | 
    Rejected: {summary.rejected_decisions} | 
    Archived: {summary.archived_decisions}
    """
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Data table
    if report.data:
        table_data = [
            ["Decision ID", "Title", "Category", "Status", "Created By", "Created Date", "Alternatives", "Approvals", "Tags"]
        ]
        
        for row in report.data:
            table_data.append([
                str(row.decision_id),
                row.decision_title[:30],  # Truncate long titles
                row.category,
                row.status,
                row.created_by,
                row.created_date.strftime("%Y-%m-%d"),
                str(row.number_of_alternatives),
                str(row.number_of_approvals),
                ", ".join(row.tags) if row.tags else ""
            ])
        
        table = Table(table_data, colWidths=[0.6*inch, 1.2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.8*inch, 0.8*inch, 1.0*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_decisions_excel(report):
    """Generate Excel for decision reports"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Decisions"
    
    # Headers
    headers = ["Decision ID", "Title", "Category", "Status", "Created By", "Created Date", "Updated Date", "Alternatives", "Approvals", "Tags"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row in report.data:
        ws.append([
            row.decision_id,
            row.decision_title,
            row.category,
            row.status,
            row.created_by,
            row.created_date.strftime("%Y-%m-%d"),
            row.updated_date.strftime("%Y-%m-%d"),
            row.number_of_alternatives,
            row.number_of_approvals,
            ", ".join(row.tags) if row.tags else ""
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 25
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_approvals_pdf(report, approval_status, start_date, end_date):
    """Generate PDF for approval reports"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph("Approval Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Generated date and filters
    generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"<b>Generated Date:</b> {generated_date}", styles['Normal']))
    
    filter_text = "Filters: "
    if approval_status:
        filter_text += f"Status: {approval_status}, "
    if start_date:
        filter_text += f"Start Date: {start_date.strftime('%Y-%m-%d')}, "
    if end_date:
        filter_text += f"End Date: {end_date.strftime('%Y-%m-%d')}, "
    
    if filter_text == "Filters: ":
        filter_text += "None"
    else:
        filter_text = filter_text.rstrip(", ")
    
    elements.append(Paragraph(f"<b>{filter_text}</b>", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Summary
    summary = report.summary
    summary_text = f"""
    <b>Summary Statistics:</b><br/>
    Total Approvals: {summary.total_approvals} | 
    Pending: {summary.pending_approvals} | 
    Approved: {summary.approved_approvals} | 
    Rejected: {summary.rejected_approvals}<br/>
    Average Turnaround Time: {summary.average_approval_turnaround_time_hours:.2f} hours | 
    Completion Rate: {summary.approval_completion_rate:.2f}%
    """
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Data table
    if report.data:
        table_data = [
            ["Approval ID", "Decision ID", "Decision Title", "Reviewer", "Level", "Status", "Assigned", "Completed", "Turnaround (hrs)"]
        ]
        
        for row in report.data:
            completed = row.completed_date.strftime("%Y-%m-%d") if row.completed_date else "N/A"
            turnaround = f"{row.approval_turnaround_time_hours:.2f}" if row.approval_turnaround_time_hours else "N/A"
            
            table_data.append([
                str(row.approval_id),
                str(row.decision_id),
                row.decision_title[:25],
                row.reviewer,
                str(row.approval_level),
                row.approval_status,
                row.assigned_date.strftime("%Y-%m-%d"),
                completed,
                turnaround
            ])
        
        table = Table(table_data, colWidths=[0.7*inch, 0.8*inch, 1.2*inch, 0.9*inch, 0.6*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1.0*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_approvals_excel(report):
    """Generate Excel for approval reports"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Approvals"
    
    # Headers
    headers = ["Approval ID", "Decision ID", "Decision Title", "Reviewer", "Level", "Status", "Assigned Date", "Completed Date", "Turnaround (hrs)"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row in report.data:
        completed = row.completed_date.strftime("%Y-%m-%d") if row.completed_date else "N/A"
        turnaround = f"{row.approval_turnaround_time_hours:.2f}" if row.approval_turnaround_time_hours else "N/A"
        
        ws.append([
            row.approval_id,
            row.decision_id,
            row.decision_title,
            row.reviewer,
            row.approval_level,
            row.approval_status,
            row.assigned_date.strftime("%Y-%m-%d"),
            completed,
            turnaround
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_teams_pdf(report, team_name, decision_status, category, start_date, end_date):
    """Generate PDF for team reports"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph("Team Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Generated date and filters
    generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"<b>Generated Date:</b> {generated_date}", styles['Normal']))
    
    filter_text = "Filters: "
    if team_name:
        filter_text += f"Team: {team_name}, "
    if decision_status:
        filter_text += f"Status: {decision_status}, "
    if category:
        filter_text += f"Category: {category}, "
    if start_date:
        filter_text += f"Start Date: {start_date.strftime('%Y-%m-%d')}, "
    if end_date:
        filter_text += f"End Date: {end_date.strftime('%Y-%m-%d')}, "
    
    if filter_text == "Filters: ":
        filter_text += "None"
    else:
        filter_text = filter_text.rstrip(", ")
    
    elements.append(Paragraph(f"<b>{filter_text}</b>", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Data table
    if report.data:
        table_data = [
            ["Team Name", "Members", "Total Decisions", "Approved", "Rejected", "Pending", "Draft", "Avg Turnaround (hrs)"]
        ]
        
        for row in report.data:
            table_data.append([
                row.team_name,
                str(row.number_of_members),
                str(row.decision_stats.total_decisions),
                str(row.decision_stats.approved_decisions),
                str(row.decision_stats.rejected_decisions),
                str(row.decision_stats.pending_decisions),
                str(row.decision_stats.draft_decisions),
                f"{row.approval_stats.average_turnaround_time_hours:.2f}"
            ])
        
        table = Table(table_data, colWidths=[1.5*inch, 0.8*inch, 1.2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.8*inch, 1.2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_teams_excel(report):
    """Generate Excel for team reports"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Teams"
    
    # Headers
    headers = ["Team Name", "Members", "Total Decisions", "Approved", "Rejected", "Pending", "Draft", "Avg Turnaround (hrs)"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row in report.data:
        ws.append([
            row.team_name,
            row.number_of_members,
            row.decision_stats.total_decisions,
            row.decision_stats.approved_decisions,
            row.decision_stats.rejected_decisions,
            row.decision_stats.pending_decisions,
            row.decision_stats.draft_decisions,
            f"{row.approval_stats.average_turnaround_time_hours:.2f}"
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 18
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def generate_audit_pdf(report, user_id, action, entity_type, start_date, end_date):
    """Generate PDF for audit reports"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter), topMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1f4788'),
        spaceAfter=12,
        alignment=1  # Center
    )
    elements.append(Paragraph("Audit Report", title_style))
    elements.append(Spacer(1, 12))
    
    # Generated date and filters
    generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    elements.append(Paragraph(f"<b>Generated Date:</b> {generated_date}", styles['Normal']))
    
    filter_text = "Filters: "
    if user_id:
        filter_text += f"User ID: {user_id}, "
    if action:
        filter_text += f"Action: {action}, "
    if entity_type:
        filter_text += f"Entity Type: {entity_type}, "
    if start_date:
        filter_text += f"Start Date: {start_date.strftime('%Y-%m-%d')}, "
    if end_date:
        filter_text += f"End Date: {end_date.strftime('%Y-%m-%d')}, "
    
    if filter_text == "Filters: ":
        filter_text += "None"
    else:
        filter_text = filter_text.rstrip(", ")
    
    elements.append(Paragraph(f"<b>{filter_text}</b>", styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Data table
    if report.data:
        table_data = [
            ["Audit ID", "User", "Action", "Entity Type", "Entity ID", "Description", "Timestamp"]
        ]
        
        for row in report.data:
            table_data.append([
                str(row.audit_id),
                row.user,
                row.action,
                row.entity_type,
                str(row.entity_id) if row.entity_id else "N/A",
                row.description[:40],  # Truncate long descriptions
                row.timestamp.strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        table = Table(table_data, colWidths=[0.7*inch, 1.0*inch, 1.0*inch, 1.2*inch, 0.8*inch, 1.8*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f4788')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')]),
        ]))
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer


def generate_audit_excel(report):
    """Generate Excel for audit reports"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"
    
    # Headers
    headers = ["Audit ID", "User", "Action", "Entity Type", "Entity ID", "Description", "Timestamp"]
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1f4788", end_color="1f4788", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    for row in report.data:
        ws.append([
            row.audit_id,
            row.user,
            row.action,
            row.entity_type,
            row.entity_id if row.entity_id else "N/A",
            row.description,
            row.timestamp.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 20
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
