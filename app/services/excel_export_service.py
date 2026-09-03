import io
from datetime import datetime
from typing import Any, Dict, List, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _create_styled_workbook(
    report_title: str,
    filters: Dict[str, Any],
    summary: Dict[str, Any],
    data_sheet_title: str,
    headers: List[str],
    rows: List[List[Any]],
) -> bytes:
    wb = openpyxl.Workbook()
    
    # Styles
    title_font = Font(name="Calibri", size=16, bold=True, color="1E293B")
    subtitle_font = Font(name="Calibri", size=10, italic=True, color="64748B")
    section_font = Font(name="Calibri", size=12, bold=True, color="0F172A")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True, color="1E293B")
    regular_font = Font(name="Calibri", size=10, color="334155")

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    summary_header_fill = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    meta_box_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

    # -------------------------------------------------------------
    # Sheet 1: Summary & Filters
    # -------------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary & Filters"
    ws_summary.views.sheetView[0].showGridLines = True

    ws_summary.cell(row=1, column=1, value=report_title).font = title_font
    ws_summary.cell(
        row=2,
        column=1,
        value=f"Generated on {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')} • Expert Decision Replay Platform",
    ).font = subtitle_font

    # Filters section
    ws_summary.cell(row=4, column=1, value="Applied Query Filters").font = section_font
    ws_summary.cell(row=5, column=1, value="Filter Parameter").font = header_font
    ws_summary.cell(row=5, column=1).fill = summary_header_fill
    ws_summary.cell(row=5, column=2, value="Applied Value").font = header_font
    ws_summary.cell(row=5, column=2).fill = summary_header_fill
    ws_summary.cell(row=5, column=1).border = cell_border
    ws_summary.cell(row=5, column=2).border = cell_border

    curr_row = 6
    active_filters = {k: v for k, v in filters.items() if v is not None and v != ""}
    if not active_filters:
        ws_summary.cell(row=curr_row, column=1, value="Filters").font = regular_font
        ws_summary.cell(row=curr_row, column=2, value="All records included (no filter)").font = regular_font
        ws_summary.cell(row=curr_row, column=1).border = cell_border
        ws_summary.cell(row=curr_row, column=2).border = cell_border
        curr_row += 1
    else:
        for k, v in active_filters.items():
            val_str = ", ".join(str(x) for x in v) if isinstance(v, list) else str(v)
            ws_summary.cell(row=curr_row, column=1, value=k.replace('_', ' ').title()).font = bold_font
            ws_summary.cell(row=curr_row, column=2, value=val_str).font = regular_font
            ws_summary.cell(row=curr_row, column=1).border = cell_border
            ws_summary.cell(row=curr_row, column=2).border = cell_border
            curr_row += 1

    # Summary metrics section
    curr_row += 2
    ws_summary.cell(row=curr_row, column=1, value="Summary Statistics").font = section_font
    curr_row += 1
    ws_summary.cell(row=curr_row, column=1, value="Metric").font = header_font
    ws_summary.cell(row=curr_row, column=1).fill = summary_header_fill
    ws_summary.cell(row=curr_row, column=2, value="Value").font = header_font
    ws_summary.cell(row=curr_row, column=2).fill = summary_header_fill
    ws_summary.cell(row=curr_row, column=1).border = cell_border
    ws_summary.cell(row=curr_row, column=2).border = cell_border
    curr_row += 1

    for k, v in summary.items():
        if isinstance(v, dict):
            val_display = ", ".join(f"{dk}: {dv}" for dk, dv in v.items())
        elif isinstance(v, float):
            val_display = f"{v:.2f}"
        else:
            val_display = str(v)

        ws_summary.cell(row=curr_row, column=1, value=k.replace('_', ' ').title()).font = bold_font
        ws_summary.cell(row=curr_row, column=2, value=val_display).font = regular_font
        ws_summary.cell(row=curr_row, column=1).border = cell_border
        ws_summary.cell(row=curr_row, column=2).border = cell_border
        curr_row += 1

    # Auto-width for summary sheet
    for col in ws_summary.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_summary.column_dimensions[col_letter].width = max(max_len + 4, 25)

    # -------------------------------------------------------------
    # Sheet 2: Data Records
    # -------------------------------------------------------------
    ws_data = wb.create_sheet(title=data_sheet_title)
    ws_data.views.sheetView[0].showGridLines = True

    # Header row
    for col_idx, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border
    ws_data.row_dimensions[1].height = 25

    # Data rows
    for row_idx, r in enumerate(rows, 2):
        row_fill = zebra_fill if row_idx % 2 == 0 else white_fill
        for col_idx, val in enumerate(r, 1):
            cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
            cell.font = regular_font
            cell.fill = row_fill
            cell.border = cell_border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif isinstance(val, str) and (val in ["Draft", "Under Review", "Approved", "Rejected", "Archived", "Pending"]):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_data.row_dimensions[row_idx].height = 20

    # Auto-adjust column widths
    for col in ws_data.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = get_column_letter(col[0].column)
        ws_data.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# =============================================================================
# 1. DECISIONS REPORT EXCEL
# =============================================================================

def generate_decision_report_excel(
    items: List[dict],
    summary: dict,
    filters: Dict[str, Any],
) -> bytes:
    headers = [
        "Decision ID",
        "Title",
        "Category",
        "Status",
        "Created By (ID)",
        "Created By (Name)",
        "Created Date",
        "Updated Date",
        "Alternatives Count",
        "Approvals Count",
        "Tags",
    ]

    rows = []
    for it in items:
        tags_str = ", ".join(it.get('tags', [])) if it.get('tags') else ""
        created_str = it['created_at'].strftime('%Y-%m-%d %H:%M') if isinstance(it.get('created_at'), datetime) else str(it.get('created_at', ''))[:16]
        updated_str = it['updated_at'].strftime('%Y-%m-%d %H:%M') if isinstance(it.get('updated_at'), datetime) else str(it.get('updated_at', ''))[:16]

        rows.append([
            it.get('decision_id'),
            it.get('title'),
            it.get('category'),
            it.get('status'),
            it.get('created_by'),
            it.get('created_by_name') or "",
            created_str,
            updated_str,
            it.get('number_of_alternatives', 0),
            it.get('number_of_approvals', 0),
            tags_str,
        ])

    return _create_styled_workbook(
        report_title="Decisions Detailed Report",
        filters=filters,
        summary=summary,
        data_sheet_title="Decisions",
        headers=headers,
        rows=rows,
    )


# =============================================================================
# 2. APPROVALS REPORT EXCEL
# =============================================================================

def generate_approval_report_excel(
    items: List[dict],
    summary: dict,
    filters: Dict[str, Any],
) -> bytes:
    headers = [
        "Approval ID",
        "Decision ID",
        "Decision Title",
        "Reviewer ID",
        "Reviewer Name",
        "Reviewer Email",
        "Approval Level",
        "Approval Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround Time (Hours)",
    ]

    rows = []
    for it in items:
        assigned_str = it['assigned_date'].strftime('%Y-%m-%d %H:%M') if isinstance(it.get('assigned_date'), datetime) else str(it.get('assigned_date', ''))[:16]
        completed_str = it['completed_date'].strftime('%Y-%m-%d %H:%M') if isinstance(it.get('completed_date'), datetime) else (str(it.get('completed_date', ''))[:16] if it.get('completed_date') else "")
        tt = it.get('turnaround_time_hours')

        rows.append([
            it.get('approval_id'),
            it.get('decision_id'),
            it.get('decision_title'),
            it.get('reviewer_id'),
            it.get('reviewer_name') or "",
            it.get('reviewer_email') or "",
            it.get('approval_level', 1),
            it.get('approval_status'),
            assigned_str,
            completed_str,
            round(tt, 2) if tt is not None else "",
        ])

    return _create_styled_workbook(
        report_title="Approvals Workflow & Performance Report",
        filters=filters,
        summary=summary,
        data_sheet_title="Approvals",
        headers=headers,
        rows=rows,
    )


# =============================================================================
# 3. TEAMS REPORT EXCEL
# =============================================================================

def generate_team_report_excel(
    items: List[dict],
    summary: dict,
    filters: Dict[str, Any],
) -> bytes:
    headers = [
        "Team / Department Name",
        "Number of Members",
        "Total Decisions",
        "Approved Decisions",
        "Rejected Decisions",
        "Pending Decisions",
        "Total Approvals",
        "Approved Approvals",
        "Rejected Approvals",
        "Pending Approvals",
        "Avg Turnaround Time (Hours)",
    ]

    rows = []
    for it in items:
        app_stats = it.get('team_approval_statistics', {})
        avg_tt = app_stats.get('average_turnaround_time_hours')

        rows.append([
            it.get('team_name', 'General'),
            it.get('number_of_members', 0),
            it.get('total_decisions', 0),
            it.get('approved_decisions', 0),
            it.get('rejected_decisions', 0),
            it.get('pending_decisions', 0),
            app_stats.get('total_approvals', 0),
            app_stats.get('approved_approvals', 0),
            app_stats.get('rejected_approvals', 0),
            app_stats.get('pending_approvals', 0),
            round(avg_tt, 2) if avg_tt is not None else "",
        ])

    return _create_styled_workbook(
        report_title="Team Activity & Approval Statistics Report",
        filters=filters,
        summary=summary,
        data_sheet_title="Teams",
        headers=headers,
        rows=rows,
    )


# =============================================================================
# 4. AUDIT REPORT EXCEL
# =============================================================================

def generate_audit_report_excel(
    items: List[dict],
    summary: dict,
    filters: Dict[str, Any],
) -> bytes:
    headers = [
        "Audit ID",
        "Timestamp",
        "User ID",
        "User Name",
        "User Email",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "IP Address",
        "Request Method",
        "Endpoint",
    ]

    rows = []
    for it in items:
        ts_str = it['timestamp'].strftime('%Y-%m-%d %H:%M:%S') if isinstance(it.get('timestamp'), datetime) else str(it.get('timestamp', ''))[:19]

        rows.append([
            it.get('id'),
            ts_str,
            it.get('user_id') or "",
            it.get('user_name') or "",
            it.get('user_email') or "",
            it.get('action'),
            it.get('entity_type'),
            it.get('entity_id') or "",
            it.get('description'),
            it.get('ip_address') or "",
            it.get('request_method') or "",
            it.get('endpoint') or "",
        ])

    return _create_styled_workbook(
        report_title="System Audit & Compliance Activity Report",
        filters=filters,
        summary=summary,
        data_sheet_title="Audit Trail",
        headers=headers,
        rows=rows,
    )
