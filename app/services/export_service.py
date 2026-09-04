import io
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def generate_pdf(title: str, headers: list, rows: list, summary: dict = None) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph(title, styles["Title"]))
    elements.append(Paragraph(
        f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 12))

    # Summary
    if summary:
        elements.append(Paragraph("Summary", styles["Heading2"]))
        for key, value in summary.items():
            elements.append(Paragraph(f"{key}: {value}", styles["Normal"]))
        elements.append(Spacer(1, 12))

    # Table
    if rows:
        table_data = [headers] + rows
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2C3E50")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#ECF0F1")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("PADDING", (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
    else:
        elements.append(Paragraph("No data found.", styles["Normal"]))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()



def generate_excel(title: str, headers: list, rows: list, summary: dict = None) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel sheet name limit

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    summary_font = Font(bold=True)

    row_num = 1

    # Title
    ws.cell(row=row_num, column=1, value=title).font = Font(bold=True, size=14)
    row_num += 1
    ws.cell(row=row_num, column=1, value=f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    row_num += 2

    # Summary
    if summary:
        ws.cell(row=row_num, column=1, value="Summary").font = summary_font
        row_num += 1
        for key, value in summary.items():
            ws.cell(row=row_num, column=1, value=str(key))
            ws.cell(row=row_num, column=2, value=str(value))
            row_num += 1
        row_num += 1

    # Headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    row_num += 1

    # Data rows
    for row in rows:
        for col_num, value in enumerate(row, 1):
            ws.cell(row=row_num, column=col_num, value=str(value) if value is not None else "")
        row_num += 1

    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()