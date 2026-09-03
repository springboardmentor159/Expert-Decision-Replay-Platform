from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


def _format_value(value: Any) -> str:
    if value is None:
        return ""

    if hasattr(value, "isoformat"):
        return value.isoformat()

    return str(value)


def generate_excel(
    title: str,
    columns: list[str],
    rows: list[dict],
    filters: dict | None = None,
    summary: dict | None = None,
) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    # Title
    worksheet["A1"] = title
    worksheet["A1"].font = Font(bold=True, size=16)

    current_row = 3

    # Generated/filter information
    if filters:
        worksheet.cell(
            row=current_row,
            column=1,
            value="Filters",
        ).font = Font(bold=True)

        current_row += 1

        for key, value in filters.items():
            worksheet.cell(
                row=current_row,
                column=1,
                value=key,
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=_format_value(value),
            )
            current_row += 1

        current_row += 1

    # Summary
    if summary:
        worksheet.cell(
            row=current_row,
            column=1,
            value="Summary",
        ).font = Font(bold=True)

        current_row += 1

        for key, value in summary.items():
            worksheet.cell(
                row=current_row,
                column=1,
                value=key,
            )
            worksheet.cell(
                row=current_row,
                column=2,
                value=_format_value(value),
            )
            current_row += 1

        current_row += 1

    # Column headers
    for column_index, column_name in enumerate(
        columns,
        start=1,
    ):
        cell = worksheet.cell(
            row=current_row,
            column=column_index,
            value=column_name,
        )
        cell.font = Font(bold=True)

    current_row += 1

    # Data
    for row in rows:
        for column_index, column_name in enumerate(
            columns,
            start=1,
        ):
            worksheet.cell(
                row=current_row,
                column=column_index,
                value=_format_value(
                    row.get(column_name)
                ),
            )

        current_row += 1

    # Automatic column widths
    for column_cells in worksheet.columns:
        max_length = 0

        for cell in column_cells:
            value = _format_value(cell.value)
            max_length = max(
                max_length,
                len(value),
            )

        worksheet.column_dimensions[
            column_cells[0].column_letter
        ].width = min(
            max(max_length + 2, 12),
            50,
        )

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def generate_pdf(
    title: str,
    columns: list[str],
    rows: list[dict],
    filters: dict | None = None,
    summary: dict | None = None,
) -> BytesIO:
    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=10 * mm,
        leftMargin=10 * mm,
        topMargin=10 * mm,
        bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()

    story = []

    story.append(
        Paragraph(
            title,
            styles["Title"],
        )
    )

    story.append(
        Spacer(
            1,
            8,
        )
    )

    if filters:
        filter_text = "<b>Filters:</b> "

        filter_parts = []

        for key, value in filters.items():
            if value is not None:
                filter_parts.append(
                    f"{key}: {_format_value(value)}"
                )

        filter_text += ", ".join(
            filter_parts
        )

        story.append(
            Paragraph(
                filter_text,
                styles["Normal"],
            )
        )

        story.append(
            Spacer(
                1,
                6,
            )
        )

    if summary:
        summary_text = "<b>Summary:</b> "

        summary_parts = []

        for key, value in summary.items():
            summary_parts.append(
                f"{key}: {_format_value(value)}"
            )

        summary_text += ", ".join(
            summary_parts
        )

        story.append(
            Paragraph(
                summary_text,
                styles["Normal"],
            )
        )

        story.append(
            Spacer(
                1,
                8,
            )
        )

    table_data = [
        columns
    ]

    for row in rows:
        table_data.append(
            [
                _format_value(
                    row.get(column)
                )
                for column in columns
            ]
        )

    if len(table_data) == 1:
        table_data.append(
            [
                "No matching records"
            ]
            + [""] * (len(columns) - 1)
        )

    table = Table(
        table_data,
        repeatRows=1,
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.lightgrey,
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.black,
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold",
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    7,
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey,
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP",
                ),
                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    4,
                ),
            ]
        )
    )

    story.append(table)

    document.build(story)

    output.seek(0)

    return output