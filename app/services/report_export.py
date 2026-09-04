from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)


def generate_excel(
    title: str,
    columns: list[str],
    rows: list[list]
):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet["A1"] = title
    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A2"] = "Generated At"
    worksheet["B2"] = datetime.utcnow().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    header_row = 4

    for column_index, column_name in enumerate(columns, start=1):
        cell = worksheet.cell(
            row=header_row,
            column=column_index
        )
        cell.value = column_name
        cell.font = Font(bold=True)
        cell.alignment = Alignment(
            horizontal="center"
        )

    for row_index, row in enumerate(rows, start=header_row + 1):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value
            )

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 40)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output


def generate_pdf(
    title: str,
    columns: list[str],
    rows: list[list],
    summary: list[str] | None = None
):
    output = BytesIO()

    document = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            title,
            styles["Title"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(
        Paragraph(
            f"Generated At: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    if summary:
        for item in summary:
            elements.append(
                Paragraph(
                    item,
                    styles["Normal"]
                )
            )

        elements.append(
            Spacer(1, 10)
        )

    table_data = [
        columns
    ]

    for row in rows:
        table_data.append(
            [
                "" if value is None else str(value)
                for value in row
            ]
        )

    table = Table(
        table_data,
        repeatRows=1
    )

    table.setStyle(
        TableStyle(
            [
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.black
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),
                (
                    "FONTSIZE",
                    (0, 0),
                    (-1, -1),
                    7
                ),
            ]
        )
    )

    elements.append(table)

    document.build(elements)

    output.seek(0)

    return output