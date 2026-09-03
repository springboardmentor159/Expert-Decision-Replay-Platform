from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# =========================================================
# COMMON HELPERS
# =========================================================

HEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="1F4E78",
)

SUBHEADER_FILL = PatternFill(
    fill_type="solid",
    fgColor="5B9BD5",
)

WHITE_FONT = Font(
    color="FFFFFF",
    bold=True,
)

TITLE_FONT = Font(
    bold=True,
    size=16,
)

BOLD_FONT = Font(
    bold=True,
)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def format_value(value):
    if value is None:
        return "-"

    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")

    return value


def create_workbook(title: str):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = title[:31]

    return workbook, worksheet


def add_report_header(
    worksheet,
    title: str,
    filters: dict,
    summary: dict,
):
    worksheet["A1"] = "Expert Decision Replay Platform"
    worksheet["A1"].font = TITLE_FONT

    worksheet["A2"] = title
    worksheet["A2"].font = Font(
        bold=True,
        size=13,
    )

    worksheet["A3"] = (
        "Generated: "
        + datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    )

    row = 5

    # -----------------------------------------------------
    # FILTERS
    # -----------------------------------------------------

    worksheet.cell(
        row=row,
        column=1,
        value="Applied Filters",
    )
    worksheet.cell(
        row=row,
        column=1,
    ).font = BOLD_FONT

    row += 1

    worksheet.cell(
        row=row,
        column=1,
        value="Filter",
    )
    worksheet.cell(
        row=row,
        column=2,
        value="Value",
    )

    for cell in worksheet[row]:
        cell.fill = SUBHEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    row += 1

    applied_filters = {
        key: value
        for key, value in filters.items()
        if value is not None and value != ""
    }

    if applied_filters:
        for key, value in applied_filters.items():
            readable_key = key.replace("_", " ").title()

            worksheet.cell(
                row=row,
                column=1,
                value=readable_key,
            )

            worksheet.cell(
                row=row,
                column=2,
                value=format_value(value),
            )

            worksheet.cell(
                row=row,
                column=1,
            ).border = THIN_BORDER

            worksheet.cell(
                row=row,
                column=2,
            ).border = THIN_BORDER

            row += 1
    else:
        worksheet.cell(
            row=row,
            column=1,
            value="No filters applied",
        )
        row += 1

    # -----------------------------------------------------
    # SUMMARY
    # -----------------------------------------------------

    row += 1

    worksheet.cell(
        row=row,
        column=1,
        value="Summary",
    )
    worksheet.cell(
        row=row,
        column=1,
    ).font = BOLD_FONT

    row += 1

    worksheet.cell(
        row=row,
        column=1,
        value="Metric",
    )
    worksheet.cell(
        row=row,
        column=2,
        value="Value",
    )

    for cell in worksheet[row]:
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.border = THIN_BORDER

    row += 1

    for key, value in summary.items():
        readable_key = key.replace("_", " ").title()

        if isinstance(value, float):
            value = round(value, 2)

        worksheet.cell(
            row=row,
            column=1,
            value=readable_key,
        )

        worksheet.cell(
            row=row,
            column=2,
            value=format_value(value),
        )

        worksheet.cell(
            row=row,
            column=1,
        ).border = THIN_BORDER

        worksheet.cell(
            row=row,
            column=2,
        ).border = THIN_BORDER

        row += 1

    return row + 2


def add_data_table(
    worksheet,
    start_row: int,
    headers: list,
    rows: list,
):
    # -----------------------------------------------------
    # HEADERS
    # -----------------------------------------------------

    for column_index, header in enumerate(
        headers,
        start=1,
    ):
        cell = worksheet.cell(
            row=start_row,
            column=column_index,
            value=header,
        )

        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )
        cell.border = THIN_BORDER

    # -----------------------------------------------------
    # DATA
    # -----------------------------------------------------

    for row_index, row_data in enumerate(
        rows,
        start=start_row + 1,
    ):
        for column_index, value in enumerate(
            row_data,
            start=1,
        ):
            cell = worksheet.cell(
                row=row_index,
                column=column_index,
                value=format_value(value),
            )

            cell.border = THIN_BORDER
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    # -----------------------------------------------------
    # FILTER
    # -----------------------------------------------------

    if rows:
        last_row = start_row + len(rows)
        last_column = len(headers)

        worksheet.auto_filter.ref = (
            f"A{start_row}:"
            f"{get_column_letter(last_column)}{last_row}"
        )

    worksheet.freeze_panes = f"A{start_row + 1}"

    # -----------------------------------------------------
    # COLUMN WIDTHS
    # -----------------------------------------------------

    for column_index in range(
        1,
        len(headers) + 1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        max_length = len(
            str(
                worksheet.cell(
                    row=start_row,
                    column=column_index,
                ).value
            )
        )

        for row_index in range(
            start_row + 1,
            worksheet.max_row + 1,
        ):
            value = worksheet.cell(
                row=row_index,
                column=column_index,
            ).value

            if value is not None:
                max_length = max(
                    max_length,
                    len(str(value)),
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max(max_length + 2, 12),
            45,
        )


def save_workbook(workbook):
    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer


# =========================================================
# DECISION EXCEL
# =========================================================

def generate_decision_excel(
    report,
    filters,
):
    workbook, worksheet = create_workbook(
        "Decision Report"
    )

    start_row = add_report_header(
        worksheet=worksheet,
        title="Decision Report",
        filters=filters,
        summary=report.get("summary", {}),
    )

    headers = [
        "Decision ID",
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date",
        "Updated Date",
        "Alternatives",
        "Approvals",
        "Tags",
    ]

    rows = []

    for item in report.get("data", []):
        creator = item.get(
            "created_by",
            {},
        )

        rows.append([
            item.get("decision_id"),
            item.get("title"),
            item.get("category"),
            item.get("status"),
            creator.get("name"),
            item.get("created_date"),
            item.get("updated_date"),
            item.get("number_alternatives"),
            item.get("number_approvals"),
            ", ".join(
                item.get("tags", [])
            ) or "-",
        ])

    add_data_table(
        worksheet,
        start_row,
        headers,
        rows,
    )

    return save_workbook(workbook)


# =========================================================
# APPROVAL EXCEL
# =========================================================

def generate_approval_excel(
    report,
    filters,
):
    workbook, worksheet = create_workbook(
        "Approval Report"
    )

    start_row = add_report_header(
        worksheet=worksheet,
        title="Approval Report",
        filters=filters,
        summary=report.get("summary", {}),
    )

    headers = [
        "Approval ID",
        "Decision ID",
        "Decision Title",
        "Reviewer",
        "Approval Level",
        "Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround Hours",
    ]

    rows = []

    for item in report.get("data", []):
        reviewer = item.get(
            "reviewer",
            {},
        )

        rows.append([
            item.get("approval_id"),
            item.get("decision_id"),
            item.get("decision_title"),
            reviewer.get("name"),
            item.get("approval_level"),
            item.get("approval_status"),
            item.get("assigned_date"),
            item.get("completed_date"),
            item.get("turnaround_time_hours"),
        ])

    add_data_table(
        worksheet,
        start_row,
        headers,
        rows,
    )

    return save_workbook(workbook)


# =========================================================
# TEAM EXCEL
# =========================================================

def generate_team_excel(
    report,
    filters,
):
    workbook, worksheet = create_workbook(
        "Team Report"
    )

    start_row = add_report_header(
        worksheet=worksheet,
        title="Team Report",
        filters=filters,
        summary=report.get("summary", {}),
    )

    headers = [
        "Team",
        "Members",
        "Decisions",
        "Approved",
        "Rejected",
        "Pending",
        "Approval Total",
        "Approval Approved",
        "Approval Rejected",
        "Approval Pending",
        "Completion %",
    ]

    rows = []

    for item in report.get("data", []):
        statistics = item.get(
            "approval_statistics",
            {},
        )

        rows.append([
            item.get("team"),
            item.get("member_count"),
            item.get("total_decisions"),
            item.get("approved"),
            item.get("rejected"),
            item.get("pending"),
            statistics.get("total"),
            statistics.get("approved"),
            statistics.get("rejected"),
            statistics.get("pending"),
            statistics.get("completion_rate"),
        ])

    add_data_table(
        worksheet,
        start_row,
        headers,
        rows,
    )

    return save_workbook(workbook)


# =========================================================
# AUDIT EXCEL
# =========================================================

def generate_audit_excel(
    report,
    filters,
):
    workbook, worksheet = create_workbook(
        "Audit Report"
    )

    start_row = add_report_header(
        worksheet=worksheet,
        title="Audit Report",
        filters=filters,
        summary=report.get("summary", {}),
    )

    headers = [
        "Audit ID",
        "User",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "Timestamp",
        "IP Address",
        "Method",
        "Endpoint",
    ]

    rows = []

    for item in report.get("data", []):
        user = item.get(
            "user",
            {},
        )

        rows.append([
            item.get("audit_id"),
            user.get("name"),
            item.get("action"),
            item.get("entity_type"),
            item.get("entity_id"),
            item.get("description"),
            item.get("timestamp"),
            item.get("ip_address"),
            item.get("request_method"),
            item.get("endpoint"),
        ])

    add_data_table(
        worksheet,
        start_row,
        headers,
        rows,
    )

    return save_workbook(workbook)