import io
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.schemas.report import (
    DecisionReportItem,
    DecisionReportSummary,
    ApprovalReportItem,
    ApprovalReportSummary,
    TeamReportItem,
    TeamReportSummary,
    AuditReportItem,
    AuditReportSummary,
)

# Common styling palettes
TITLE_FILL = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
APPROVAL_HEADER_FILL = PatternFill(start_color="047857", end_color="047857", fill_type="solid")
TEAM_HEADER_FILL = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
AUDIT_HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
ZEBRA_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
KPI_FILL = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

TITLE_FONT = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name="Calibri", size=9, italic=True, color="64748B")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="1E293B")
REGULAR_FONT = Font(name="Calibri", size=10, color="1E293B")
KPI_LABEL_FONT = Font(name="Calibri", size=9, bold=True, color="475569")
KPI_VAL_FONT = Font(name="Calibri", size=12, bold=True, color="0F172A")

THIN_BORDER = Border(
    left=Side(style="thin", color="CBD5E1"),
    right=Side(style="thin", color="CBD5E1"),
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1"),
)


def _apply_metadata_and_filters(ws, title: str, filters: dict, user_name: str, max_col: int = 9):
    # Title Row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Subtitle Row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
    now_str = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    sub_cell = ws.cell(
        row=2,
        column=1,
        value=f"Expert Decision Replay Platform  |  Generated: {now_str}  |  By: {user_name}",
    )
    sub_cell.font = SUBTITLE_FONT
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 18

    # Filters Row
    filter_items = [f"{k}: {v}" for k, v in filters.items() if v is not None and v != ""]
    filter_text = "Applied Filters: " + (" | ".join(filter_items) if filter_items else "None (All records)")
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
    f_cell = ws.cell(row=3, column=1, value=filter_text)
    f_cell.font = Font(name="Calibri", size=9, bold=True, color="334155")
    f_cell.fill = KPI_FILL
    f_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 20


def _auto_fit_columns(ws, max_len_cap: int = 50):
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    lines = str(cell.value).split("\n")
                    for line in lines:
                        if len(str(line)) > max_length:
                            max_length = len(str(line))
            except Exception:
                pass
        adjusted_width = min(max(max_length + 4, 12), max_len_cap)
        ws.column_dimensions[column_letter].width = adjusted_width


# ============================================================
# 1. DECISION REPORT EXCEL GENERATOR
# ============================================================

def generate_decisions_excel(
    items: list[DecisionReportItem],
    summary: DecisionReportSummary,
    filters: dict,
    user_name: str = "Administrator",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Decisions Report"
    ws.views.sheetView[0].showGridLines = True

    _apply_metadata_and_filters(ws, "Decision Management Report", filters, user_name, max_col=9)

    # KPI Summary Cards (Rows 5 & 6)
    kpis = [
        ("Total Decisions", summary.total_decisions),
        ("Draft", summary.draft_decisions),
        ("Under Review", summary.decisions_under_review),
        ("Approved", summary.approved_decisions),
        ("Rejected", summary.rejected_decisions),
        ("Archived", summary.archived_decisions),
    ]

    ws.cell(row=5, column=1, value="SUMMARY METRICS").font = BOLD_FONT

    for idx, (label, val) in enumerate(kpis, start=1):
        c_label = ws.cell(row=6, column=idx, value=label.upper())
        c_label.font = KPI_LABEL_FONT
        c_label.fill = KPI_FILL
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        c_label.border = THIN_BORDER

        c_val = ws.cell(row=7, column=idx, value=val)
        c_val.font = KPI_VAL_FONT
        c_val.fill = KPI_FILL
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = THIN_BORDER

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24

    # Table Header (Row 9)
    headers = [
        "Decision ID",
        "Decision Title",
        "Category",
        "Status",
        "Created By",
        "Created Date",
        "Updated Date",
        "Alternatives Count",
        "Approvals Count",
        "Tags",
    ]

    ws.row_dimensions[9].height = 24
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    # Table Data
    current_row = 10
    for item in items:
        tags_str = ", ".join(item.tags) if item.tags else ""
        creator_str = item.creator_name or f"User #{item.created_by}"
        row_fill = ZEBRA_FILL if (current_row % 2 == 0) else WHITE_FILL

        row_data = [
            (item.decision_id, "center"),
            (item.decision_title, "left"),
            (item.category, "left"),
            (item.status, "center"),
            (creator_str, "left"),
            (item.created_date.strftime("%Y-%m-%d %H:%M"), "center"),
            (item.updated_date.strftime("%Y-%m-%d %H:%M"), "center"),
            (item.number_of_alternatives, "center"),
            (item.number_of_approvals, "center"),
            (tags_str, "left"),
        ]

        ws.row_dimensions[current_row].height = 20
        for col_idx, (val, align) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = THIN_BORDER

        current_row += 1

    if not items:
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=len(headers))
        empty_cell = ws.cell(row=10, column=1, value="No decision records found matching the applied filters.")
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[10].height = 24

    ws.freeze_panes = "A10"
    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================
# 2. APPROVAL REPORT EXCEL GENERATOR
# ============================================================

def generate_approvals_excel(
    items: list[ApprovalReportItem],
    summary: ApprovalReportSummary,
    filters: dict,
    user_name: str = "Administrator",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approvals Report"
    ws.views.sheetView[0].showGridLines = True

    _apply_metadata_and_filters(ws, "Approval Workflow & Compliance Report", filters, user_name, max_col=9)

    avg_turnaround = (
        f"{summary.average_approval_turnaround_time_hours:.1f} hrs"
        if summary.average_approval_turnaround_time_hours is not None
        else "N/A"
    )

    kpis = [
        ("Total Approvals", summary.total_approvals),
        ("Pending", summary.pending_approvals),
        ("Approved", summary.approved_approvals),
        ("Rejected", summary.rejected_approvals),
        ("Avg Turnaround", avg_turnaround),
        ("Completion Rate", f"{summary.approval_completion_rate:.1f}%"),
    ]

    ws.cell(row=5, column=1, value="SUMMARY METRICS").font = BOLD_FONT

    for idx, (label, val) in enumerate(kpis, start=1):
        c_label = ws.cell(row=6, column=idx, value=label.upper())
        c_label.font = KPI_LABEL_FONT
        c_label.fill = KPI_FILL
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        c_label.border = THIN_BORDER

        c_val = ws.cell(row=7, column=idx, value=val)
        c_val.font = KPI_VAL_FONT
        c_val.fill = KPI_FILL
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = THIN_BORDER

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24

    headers = [
        "Approval ID",
        "Decision ID",
        "Decision Title",
        "Reviewer",
        "Approval Level",
        "Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround (Hours)",
    ]

    ws.row_dimensions[9].height = 24
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = APPROVAL_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    current_row = 10
    for item in items:
        reviewer_str = item.reviewer_name or f"User #{item.reviewer_id}"
        completed_str = item.completed_date.strftime("%Y-%m-%d %H:%M") if item.completed_date else ""
        turnaround_val = (
            round(item.approval_turnaround_time_hours, 2)
            if item.approval_turnaround_time_hours is not None
            else ""
        )
        row_fill = ZEBRA_FILL if (current_row % 2 == 0) else WHITE_FILL

        row_data = [
            (item.approval_id, "center"),
            (item.decision_id, "center"),
            (item.decision_title, "left"),
            (reviewer_str, "left"),
            (item.approval_level, "center"),
            (item.approval_status, "center"),
            (item.assigned_date.strftime("%Y-%m-%d %H:%M"), "center"),
            (completed_str, "center"),
            (turnaround_val, "center"),
        ]

        ws.row_dimensions[current_row].height = 20
        for col_idx, (val, align) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = THIN_BORDER

        current_row += 1

    if not items:
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=len(headers))
        empty_cell = ws.cell(row=10, column=1, value="No approval records found matching the applied filters.")
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[10].height = 24

    ws.freeze_panes = "A10"
    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================
# 3. TEAM REPORT EXCEL GENERATOR
# ============================================================

def generate_teams_excel(
    items: list[TeamReportItem],
    summary: TeamReportSummary,
    filters: dict,
    user_name: str = "Administrator",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Team Performance Report"
    ws.views.sheetView[0].showGridLines = True

    _apply_metadata_and_filters(ws, "Team Performance & Decision Report", filters, user_name, max_col=9)

    kpis = [
        ("Total Teams", summary.total_teams),
        ("Total Members", summary.total_members),
        ("Total Decisions", summary.total_decisions),
        ("Total Approvals", summary.total_approvals),
    ]

    ws.cell(row=5, column=1, value="OVERALL STATISTICS").font = BOLD_FONT

    for idx, (label, val) in enumerate(kpis, start=1):
        c_label = ws.cell(row=6, column=idx, value=label.upper())
        c_label.font = KPI_LABEL_FONT
        c_label.fill = KPI_FILL
        c_label.alignment = Alignment(horizontal="center", vertical="center")
        c_label.border = THIN_BORDER

        c_val = ws.cell(row=7, column=idx, value=val)
        c_val.font = KPI_VAL_FONT
        c_val.fill = KPI_FILL
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = THIN_BORDER

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24

    headers = [
        "Team / Department",
        "Members Count",
        "Total Decisions",
        "Approved Decisions",
        "Rejected Decisions",
        "Pending Decisions",
        "Total Approvals",
        "Completion Rate (%)",
        "Avg Turnaround (Hrs)",
    ]

    ws.row_dimensions[9].height = 24
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = TEAM_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    current_row = 10
    for item in items:
        avg_t = (
            round(item.team_approval_statistics.average_turnaround_time_hours, 2)
            if item.team_approval_statistics.average_turnaround_time_hours is not None
            else "N/A"
        )
        row_fill = ZEBRA_FILL if (current_row % 2 == 0) else WHITE_FILL

        row_data = [
            (item.team_name, "left"),
            (item.number_of_members, "center"),
            (item.total_decisions, "center"),
            (item.approved_decisions, "center"),
            (item.rejected_decisions, "center"),
            (item.pending_decisions, "center"),
            (item.team_approval_statistics.total_approvals, "center"),
            (f"{item.team_approval_statistics.completion_rate:.1f}%", "center"),
            (avg_t, "center"),
        ]

        ws.row_dimensions[current_row].height = 20
        for col_idx, (val, align) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = THIN_BORDER

        current_row += 1

    if not items:
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=len(headers))
        empty_cell = ws.cell(row=10, column=1, value="No team performance records found matching the applied filters.")
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[10].height = 24

    ws.freeze_panes = "A10"
    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ============================================================
# 4. AUDIT REPORT EXCEL GENERATOR
# ============================================================

def generate_audit_excel(
    items: list[AuditReportItem],
    summary: AuditReportSummary,
    filters: dict,
    user_name: str = "Administrator",
) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Activity Report"
    ws.views.sheetView[0].showGridLines = True

    _apply_metadata_and_filters(ws, "System Audit & Compliance Activity Report", filters, user_name, max_col=8)

    ws.cell(row=5, column=1, value="AUDIT SUMMARY").font = BOLD_FONT

    c_label = ws.cell(row=6, column=1, value="TOTAL RECORDS")
    c_label.font = KPI_LABEL_FONT
    c_label.fill = KPI_FILL
    c_label.alignment = Alignment(horizontal="center", vertical="center")
    c_label.border = THIN_BORDER

    c_val = ws.cell(row=7, column=1, value=summary.total_audit_records)
    c_val.font = KPI_VAL_FONT
    c_val.fill = KPI_FILL
    c_val.alignment = Alignment(horizontal="center", vertical="center")
    c_val.border = THIN_BORDER

    top_actions = ", ".join([f"{k}: {v}" for k, v in list(summary.actions_breakdown.items())[:5]]) or "None"
    ws.merge_cells(start_row=6, start_column=2, end_row=6, end_column=4)
    ws.merge_cells(start_row=7, start_column=2, end_row=7, end_column=4)
    act_label = ws.cell(row=6, column=2, value="ACTION TYPES BREAKDOWN")
    act_label.font = KPI_LABEL_FONT
    act_label.fill = KPI_FILL
    act_label.alignment = Alignment(horizontal="center", vertical="center")
    act_label.border = THIN_BORDER
    act_val = ws.cell(row=7, column=2, value=top_actions)
    act_val.font = REGULAR_FONT
    act_val.fill = KPI_FILL
    act_val.alignment = Alignment(horizontal="center", vertical="center")
    act_val.border = THIN_BORDER

    top_entities = ", ".join([f"{k}: {v}" for k, v in list(summary.entity_types_breakdown.items())[:5]]) or "None"
    ws.merge_cells(start_row=6, start_column=5, end_row=6, end_column=8)
    ws.merge_cells(start_row=7, start_column=5, end_row=7, end_column=8)
    ent_label = ws.cell(row=6, column=5, value="ENTITY TYPES BREAKDOWN")
    ent_label.font = KPI_LABEL_FONT
    ent_label.fill = KPI_FILL
    ent_label.alignment = Alignment(horizontal="center", vertical="center")
    ent_label.border = THIN_BORDER
    ent_val = ws.cell(row=7, column=5, value=top_entities)
    ent_val.font = REGULAR_FONT
    ent_val.fill = KPI_FILL
    ent_val.alignment = Alignment(horizontal="center", vertical="center")
    ent_val.border = THIN_BORDER

    ws.row_dimensions[6].height = 18
    ws.row_dimensions[7].height = 24

    headers = [
        "Audit ID",
        "Timestamp",
        "User",
        "User Email",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "IP Address",
    ]

    ws.row_dimensions[9].height = 24
    for col_idx, header_text in enumerate(headers, start=1):
        cell = ws.cell(row=9, column=col_idx, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = AUDIT_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    current_row = 10
    for item in items:
        user_str = item.user_name or f"User #{item.user_id}"
        row_fill = ZEBRA_FILL if (current_row % 2 == 0) else WHITE_FILL

        row_data = [
            (item.audit_id, "center"),
            (item.timestamp.strftime("%Y-%m-%d %H:%M:%S"), "center"),
            (user_str, "left"),
            (item.user_email or "", "left"),
            (item.action, "center"),
            (item.entity_type, "center"),
            (item.entity_id if item.entity_id is not None else "", "center"),
            (item.description, "left"),
            (item.ip_address or "", "center"),
        ]

        ws.row_dimensions[current_row].height = 20
        for col_idx, (val, align) in enumerate(row_data, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.font = REGULAR_FONT
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = THIN_BORDER

        current_row += 1

    if not items:
        ws.merge_cells(start_row=10, start_column=1, end_row=10, end_column=len(headers))
        empty_cell = ws.cell(row=10, column=1, value="No audit records found matching the applied filters.")
        empty_cell.font = Font(name="Calibri", size=10, italic=True, color="64748B")
        empty_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[10].height = 24

    ws.freeze_panes = "A10"
    _auto_fit_columns(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
