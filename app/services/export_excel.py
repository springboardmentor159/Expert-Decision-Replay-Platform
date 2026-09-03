from datetime import datetime
import io
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def _get_styles():
    navy_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    
    title_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")

    section_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True, color="0F172A")

    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    regular_font = Font(name="Calibri", size=10, color="1E293B")
    bold_font = Font(name="Calibri", size=10, bold=True, color="1E293B")

    thin_border_side = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    return {
        "navy_fill": navy_fill,
        "header_font": header_font,
        "title_fill": title_fill,
        "title_font": title_font,
        "section_fill": section_fill,
        "section_font": section_font,
        "zebra_fill": zebra_fill,
        "regular_font": regular_font,
        "bold_font": bold_font,
        "cell_border": cell_border,
    }


def _auto_fit_columns(ws, max_cols: int):
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)


# =============================================================================
# 1. DECISION REPORT EXCEL
# =============================================================================

def generate_decisions_excel(items: List[Any], summary: Any, filters_applied: Dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Decisions Report"
    styles = _get_styles()

    # Title Banner
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = "Expert Decision Replay - Decisions Report"
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata & Filters
    ws["A2"] = f"Generated On: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")
    
    active_filters = {k: v for k, v in filters_applied.items() if v is not None and v != ""}
    filter_str = ", ".join(f"{k}: {v}" for k, v in active_filters.items()) if active_filters else "None (All records)"
    ws["A3"] = f"Applied Filters: {filter_str}"
    ws["A3"].font = Font(name="Calibri", size=9, color="334155")

    # Summary Section
    ws["A5"] = "Summary Statistics"
    ws["A5"].font = styles["section_font"]
    ws["A5"].fill = styles["section_fill"]
    ws.merge_cells("A5:F5")

    sum_headers = ["Total Decisions", "Draft", "Under Review", "Approved", "Rejected", "Archived"]
    sum_values = [
        getattr(summary, "total_decisions", 0),
        getattr(summary, "draft_decisions", 0),
        getattr(summary, "under_review_decisions", 0),
        getattr(summary, "approved_decisions", 0),
        getattr(summary, "rejected_decisions", 0),
        getattr(summary, "archived_decisions", 0),
    ]
    for col_idx, (h, val) in enumerate(zip(sum_headers, sum_values), start=1):
        c_h = ws.cell(row=6, column=col_idx, value=h)
        c_h.font = styles["bold_font"]
        c_h.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_h.border = styles["cell_border"]
        c_h.alignment = Alignment(horizontal="center")

        c_v = ws.cell(row=7, column=col_idx, value=val)
        c_v.font = styles["bold_font"]
        c_v.border = styles["cell_border"]
        c_v.alignment = Alignment(horizontal="center")

    # Data Table Headers
    data_start_row = 9
    headers = [
        "Decision ID",
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date",
        "Updated Date",
        "Alternatives Count",
        "Approvals Count",
        "Tags"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=data_start_row, column=col_idx, value=h)
        c.font = styles["header_font"]
        c.fill = styles["navy_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles["cell_border"]
    ws.row_dimensions[data_start_row].height = 22

    # Data Rows
    for row_idx, item in enumerate(items, start=data_start_row + 1):
        creator = getattr(item, "creator_name", None) or f"User #{getattr(item, 'created_by', '')}"
        created_str = item.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(item, "created_at") and item.created_at else ""
        updated_str = item.updated_at.strftime("%Y-%m-%d %H:%M") if hasattr(item, "updated_at") and item.updated_at else ""
        tags_str = ", ".join(item.tags) if hasattr(item, "tags") and item.tags else ""

        row_vals = [
            item.id,
            item.title,
            item.category,
            item.status,
            creator,
            created_str,
            updated_str,
            getattr(item, "alternatives_count", 0),
            getattr(item, "approvals_count", 0),
            tags_str
        ]
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = styles["regular_font"]
            c.border = styles["cell_border"]
            if is_even:
                c.fill = styles["zebra_fill"]
            if col_idx in [1, 8, 9]:
                c.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws, len(headers))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# 2. APPROVAL REPORT EXCEL
# =============================================================================

def generate_approvals_excel(items: List[Any], summary: Any, filters_applied: Dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approvals Report"
    styles = _get_styles()

    # Title Banner
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Expert Decision Replay - Approvals Report"
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata & Filters
    ws["A2"] = f"Generated On: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

    active_filters = {k: v for k, v in filters_applied.items() if v is not None and v != ""}
    filter_str = ", ".join(f"{k}: {v}" for k, v in active_filters.items()) if active_filters else "None (All records)"
    ws["A3"] = f"Applied Filters: {filter_str}"
    ws["A3"].font = Font(name="Calibri", size=9, color="334155")

    # Summary Section
    ws["A5"] = "Summary Statistics"
    ws["A5"].font = styles["section_font"]
    ws["A5"].fill = styles["section_fill"]
    ws.merge_cells("A5:F5")

    sum_headers = ["Total Approvals", "Pending", "Approved", "Rejected", "Avg Turnaround (hrs)", "Completion Rate (%)"]
    sum_values = [
        getattr(summary, "total_approvals", 0),
        getattr(summary, "pending_approvals", 0),
        getattr(summary, "approved_approvals", 0),
        getattr(summary, "rejected_approvals", 0),
        getattr(summary, "average_turnaround_time_hours", "N/A"),
        f"{getattr(summary, 'completion_rate', 0.0)}%",
    ]
    for col_idx, (h, val) in enumerate(zip(sum_headers, sum_values), start=1):
        c_h = ws.cell(row=6, column=col_idx, value=h)
        c_h.font = styles["bold_font"]
        c_h.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_h.border = styles["cell_border"]
        c_h.alignment = Alignment(horizontal="center")

        c_v = ws.cell(row=7, column=col_idx, value=val)
        c_v.font = styles["bold_font"]
        c_v.border = styles["cell_border"]
        c_v.alignment = Alignment(horizontal="center")

    # Data Table Headers
    data_start_row = 9
    headers = [
        "Approval ID",
        "Decision ID",
        "Decision Title",
        "Reviewer",
        "Approval Level",
        "Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround (hrs)",
        "Comments"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=data_start_row, column=col_idx, value=h)
        c.font = styles["header_font"]
        c.fill = styles["navy_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles["cell_border"]
    ws.row_dimensions[data_start_row].height = 22

    # Data Rows
    for row_idx, item in enumerate(items, start=data_start_row + 1):
        reviewer = getattr(item, "reviewer_name", None) or f"User #{getattr(item, 'reviewer_id', '')}"
        assigned_str = item.created_at.strftime("%Y-%m-%d %H:%M") if hasattr(item, "created_at") and item.created_at else ""
        completed_str = item.completed_at.strftime("%Y-%m-%d %H:%M") if hasattr(item, "completed_at") and item.completed_at else ""
        turnaround = getattr(item, "turnaround_time_hours", None) if getattr(item, "turnaround_time_hours", None) is not None else ""

        row_vals = [
            item.id,
            item.decision_id,
            getattr(item, "decision_title", "") or f"Decision #{item.decision_id}",
            reviewer,
            item.approval_level,
            item.status,
            assigned_str,
            completed_str,
            turnaround,
            getattr(item, "comments", "") or ""
        ]
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = styles["regular_font"]
            c.border = styles["cell_border"]
            if is_even:
                c.fill = styles["zebra_fill"]
            if col_idx in [1, 2, 5, 6, 9]:
                c.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws, len(headers))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# 3. TEAM REPORT EXCEL
# =============================================================================

def generate_teams_excel(items: List[Any], summary: Any, filters_applied: Dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Teams Report"
    styles = _get_styles()

    # Title Banner
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Expert Decision Replay - Teams & Departments Report"
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata & Filters
    ws["A2"] = f"Generated On: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

    active_filters = {k: v for k, v in filters_applied.items() if v is not None and v != ""}
    filter_str = ", ".join(f"{k}: {v}" for k, v in active_filters.items()) if active_filters else "None (All records)"
    ws["A3"] = f"Applied Filters: {filter_str}"
    ws["A3"].font = Font(name="Calibri", size=9, color="334155")

    # Summary Section
    ws["A5"] = "Organization Summary"
    ws["A5"].font = styles["section_font"]
    ws["A5"].fill = styles["section_fill"]
    ws.merge_cells("A5:F5")

    sum_headers = ["Total Teams", "Total Members", "Total Decisions", "Approved Decisions", "Rejected Decisions", "Pending Decisions"]
    sum_values = [
        getattr(summary, "total_teams", 0),
        getattr(summary, "total_members", 0),
        getattr(summary, "total_decisions", 0),
        getattr(summary, "approved_decisions", 0),
        getattr(summary, "rejected_decisions", 0),
        getattr(summary, "pending_decisions", 0),
    ]
    for col_idx, (h, val) in enumerate(zip(sum_headers, sum_values), start=1):
        c_h = ws.cell(row=6, column=col_idx, value=h)
        c_h.font = styles["bold_font"]
        c_h.fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        c_h.border = styles["cell_border"]
        c_h.alignment = Alignment(horizontal="center")

        c_v = ws.cell(row=7, column=col_idx, value=val)
        c_v.font = styles["bold_font"]
        c_v.border = styles["cell_border"]
        c_v.alignment = Alignment(horizontal="center")

    # Data Table Headers
    data_start_row = 9
    headers = [
        "Team / Department",
        "Members Count",
        "Total Decisions",
        "Approved Decisions",
        "Rejected Decisions",
        "Pending Decisions",
        "Draft Decisions",
        "Under Review",
        "Team Approvals",
        "Approval Completion Rate",
        "Avg Turnaround (hrs)"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=data_start_row, column=col_idx, value=h)
        c.font = styles["header_font"]
        c.fill = styles["navy_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles["cell_border"]
    ws.row_dimensions[data_start_row].height = 22

    # Data Rows
    for row_idx, item in enumerate(items, start=data_start_row + 1):
        apprv_stats = getattr(item, "team_approval_statistics", None)
        tot_apprv = getattr(apprv_stats, "total_approvals", 0) if apprv_stats else 0
        comp_rate = f"{getattr(apprv_stats, 'completion_rate', 0.0)}%" if apprv_stats else "0.0%"
        avg_turn = getattr(apprv_stats, "average_turnaround_time_hours", "N/A") if apprv_stats else "N/A"

        row_vals = [
            item.team_name,
            item.member_count,
            item.total_decisions,
            item.approved_decisions,
            item.rejected_decisions,
            item.pending_decisions,
            item.draft_decisions,
            item.under_review_decisions,
            tot_apprv,
            comp_rate,
            avg_turn
        ]
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = styles["regular_font"]
            c.border = styles["cell_border"]
            if is_even:
                c.fill = styles["zebra_fill"]
            if col_idx >= 2:
                c.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws, len(headers))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# =============================================================================
# 4. AUDIT REPORT EXCEL
# =============================================================================

def generate_audit_excel(items: List[Any], summary: Any, filters_applied: Dict[str, Any]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Report"
    styles = _get_styles()

    # Title Banner
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Expert Decision Replay - Audit & Compliance Report"
    title_cell.font = styles["title_font"]
    title_cell.fill = styles["title_fill"]
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Metadata & Filters
    ws["A2"] = f"Generated On: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}"
    ws["A2"].font = Font(name="Calibri", size=9, italic=True, color="64748B")

    active_filters = {k: v for k, v in filters_applied.items() if v is not None and v != ""}
    filter_str = ", ".join(f"{k}: {v}" for k, v in active_filters.items()) if active_filters else "None (All records)"
    ws["A3"] = f"Applied Filters: {filter_str}"
    ws["A3"].font = Font(name="Calibri", size=9, color="334155")

    # Summary Section
    ws["A5"] = "Audit Summary"
    ws["A5"].font = styles["section_font"]
    ws["A5"].fill = styles["section_fill"]
    ws.merge_cells("A5:D5")

    ws["A6"] = "Total Audit Logs"
    ws["A6"].font = styles["bold_font"]
    ws["A6"].fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    ws["A6"].border = styles["cell_border"]

    ws["B6"] = getattr(summary, "total_audit_logs", 0)
    ws["B6"].font = styles["bold_font"]
    ws["B6"].border = styles["cell_border"]

    # Data Table Headers
    data_start_row = 8
    headers = [
        "Audit ID",
        "Timestamp",
        "User",
        "Action",
        "Entity Type",
        "Entity ID",
        "IP Address",
        "Description"
    ]
    for col_idx, h in enumerate(headers, start=1):
        c = ws.cell(row=data_start_row, column=col_idx, value=h)
        c.font = styles["header_font"]
        c.fill = styles["navy_fill"]
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = styles["cell_border"]
    ws.row_dimensions[data_start_row].height = 22

    # Data Rows
    for row_idx, item in enumerate(items, start=data_start_row + 1):
        user_str = getattr(item, "user_name", None) or (f"User #{item.user_id}" if getattr(item, "user_id", None) else "System")
        time_str = item.created_at.strftime("%Y-%m-%d %H:%M:%S") if hasattr(item, "created_at") and item.created_at else ""

        row_vals = [
            item.id,
            time_str,
            user_str,
            item.action,
            item.entity_type,
            item.entity_id if item.entity_id is not None else "",
            getattr(item, "ip_address", "") or "",
            item.description
        ]
        is_even = (row_idx % 2 == 0)
        for col_idx, val in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.font = styles["regular_font"]
            c.border = styles["cell_border"]
            if is_even:
                c.fill = styles["zebra_fill"]
            if col_idx in [1, 4, 5, 6]:
                c.alignment = Alignment(horizontal="center")

    _auto_fit_columns(ws, len(headers))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
