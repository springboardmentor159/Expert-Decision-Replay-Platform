from datetime import datetime
from math import ceil
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.db.database import get_db
from app.core.security import get_current_user

from app.models.decision import Decision
from app.models.user import User
from app.models.alternative import Alternative
from app.models.approval import Approval
from app.models.tag import Tag, decision_tags
from app.models.audit_log import AuditLog

from app.services.audit import ALLOWED_ACTIONS, ALLOWED_ENTITY_TYPES

from app.schemas.report import (
    DecisionReportResponse,
    DecisionReportItem,
    DecisionReportSummary,
    ApprovalReportResponse,
    ApprovalReportItem,
    ApprovalReportSummary,
    TeamReportItem,
    TeamReportResponse,
    AuditReportItem,
    AuditReportResponse,
)

from io import BytesIO
from fastapi.responses import StreamingResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

router = APIRouter(
    prefix="/reports",
    tags=["Reports"],
)


@router.get(
    "/decisions",
    response_model=DecisionReportResponse,
)
def get_decision_report(
    category: Optional[str] = Query(
        default=None,
        description="Filter by decision category",
    ),

    status: Optional[str] = Query(
        default=None,
        description="Filter by decision status",
    ),

    created_by: Optional[int] = Query(
        default=None,
        description="Filter by creator user ID",
    ),

    start_date: Optional[datetime] = Query(
        default=None,
        description="Filter decisions created on or after this date",
    ),

    end_date: Optional[datetime] = Query(
        default=None,
        description="Filter decisions created on or before this date",
    ),

    tag: Optional[str] = Query(
        default=None,
        description="Filter by tag name",
    ),

    page: int = Query(
        default=1,
        ge=1,
        description="Page number",
    ),

    page_size: int = Query(
        default=20,
        ge=1,
        le=100,
        description="Number of records per page",
    ),

    sort_by: str = Query(
        default="created_date",
        description="Allowed: created_date, updated_date, title",
    ),

    sort_order: str = Query(
        default="desc",
        description="Allowed: asc, desc",
    ),

    current_user: User = Depends(get_current_user),

    db: Session = Depends(get_db),
):
    # ---------------------------------------------------------
    # VALIDATE STATUS
    # ---------------------------------------------------------

    allowed_statuses = {
        "Draft",
        "Under Review",
        "Approved",
        "Rejected",
        "Archived",
    }

    if status is not None and status not in allowed_statuses:
        raise HTTPException(
            status_code=422,
            detail=(
                "Invalid status. Allowed values: "
                "Draft, Under Review, Approved, Rejected, Archived"
            ),
        )

    # ---------------------------------------------------------
    # VALIDATE DATE RANGE
    # ---------------------------------------------------------

    if start_date and end_date and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be greater than end_date",
        )

    # ---------------------------------------------------------
    # VALIDATE SORTING
    # ---------------------------------------------------------

    allowed_sort_fields = {
        "created_date": Decision.created_at,
        "updated_date": Decision.updated_at,
        "title": Decision.title,
    }

    if sort_by not in allowed_sort_fields:
        raise HTTPException(
            status_code=422,
            detail=(
                "Invalid sorting field. Allowed values: "
                "created_date, updated_date, title"
            ),
        )

    sort_order = sort_order.lower()

    if sort_order not in {"asc", "desc"}:
        raise HTTPException(
            status_code=422,
            detail="sort_order must be either 'asc' or 'desc'",
        )

    # ---------------------------------------------------------
    # BASE QUERY
    # ---------------------------------------------------------

    query = db.query(Decision)

    # ---------------------------------------------------------
    # FILTERS
    # ---------------------------------------------------------

    if category:
        query = query.filter(
            Decision.category == category
        )

    if status:
        query = query.filter(
            Decision.status == status
        )

    if created_by:
        query = query.filter(
            Decision.created_by == created_by
        )

    if start_date:
        query = query.filter(
            Decision.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            Decision.created_at <= end_date
        )

    # ---------------------------------------------------------
    # TAG FILTER
    # ---------------------------------------------------------

    if tag:
        query = query.join(
            decision_tags,
            decision_tags.c.decision_id == Decision.id,
        ).join(
            Tag,
            Tag.id == decision_tags.c.tag_id,
        ).filter(
            func.lower(Tag.name) == tag.lower()
        )

    # ---------------------------------------------------------
    # TOTAL RECORDS
    # ---------------------------------------------------------

    total_records = query.distinct().count()

    # ---------------------------------------------------------
    # SUMMARY
    #
    # Summary respects the applied filters except pagination.
    # ---------------------------------------------------------

    summary_query = query.distinct()

    total_decisions = summary_query.count()

    draft_decisions = summary_query.filter(
        Decision.status == "Draft"
    ).count()

    under_review_decisions = summary_query.filter(
        Decision.status == "Under Review"
    ).count()

    approved_decisions = summary_query.filter(
        Decision.status == "Approved"
    ).count()

    rejected_decisions = summary_query.filter(
        Decision.status == "Rejected"
    ).count()

    archived_decisions = summary_query.filter(
        Decision.status == "Archived"
    ).count()

    # ---------------------------------------------------------
    # SORTING
    # ---------------------------------------------------------

    sort_column = allowed_sort_fields[sort_by]

    if sort_order == "asc":
        query = query.order_by(sort_column.asc())
    else:
        query = query.order_by(sort_column.desc())

    # ---------------------------------------------------------
    # PAGINATION
    # ---------------------------------------------------------

    offset = (page - 1) * page_size

    decisions = (
        query
        .distinct()
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # ---------------------------------------------------------
    # BUILD RESPONSE
    # ---------------------------------------------------------

    report_data = []

    for decision in decisions:

        alternative_count = (
            db.query(func.count(Alternative.id))
            .filter(
                Alternative.decision_id == decision.id
            )
            .scalar()
        )

        approval_count = (
            db.query(func.count(Approval.id))
            .filter(
                Approval.decision_id == decision.id
            )
            .scalar()
        )

        tags = [
            current_tag.name
            for current_tag in decision.tags
        ]

        report_data.append(
            DecisionReportItem(
                decision_id=decision.id,
                title=decision.title,
                category=decision.category,
                status=decision.status,
                created_by=(
                    decision.creator.full_name
                    if decision.creator
                    else None
                ),
                created_date=decision.created_at,
                updated_date=decision.updated_at,
                number_of_alternatives=alternative_count or 0,
                number_of_approvals=approval_count or 0,
                tags=tags,
            )
        )

    total_pages = (
        ceil(total_records / page_size)
        if total_records > 0
        else 0
    )

    return DecisionReportResponse(
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages,
        sort_by=sort_by,
        sort_order=sort_order,
        summary=DecisionReportSummary(
            total_decisions=total_decisions,
            draft_decisions=draft_decisions,
            under_review_decisions=under_review_decisions,
            approved_decisions=approved_decisions,
            rejected_decisions=rejected_decisions,
            archived_decisions=archived_decisions,
        ),
        data=report_data,
    )
# =========================================================
# APPROVAL REPORT
# =========================================================

@router.get(
    "/approvals",
    response_model=ApprovalReportResponse,
)
def get_approval_report(
    approval_status: Optional[str] = Query(
        default=None,
        description="Filter by approval status",
    ),

    reviewer: Optional[int] = Query(
        default=None,
        description="Filter by reviewer user ID",
    ),

    decision: Optional[int] = Query(
        default=None,
        description="Filter by decision ID",
    ),

    start_date: Optional[datetime] = Query(
        default=None,
        description="Filter approvals assigned on or after this date",
    ),

    end_date: Optional[datetime] = Query(
        default=None,
        description="Filter approvals assigned on or before this date",
    ),

    page: int = Query(
        default=1,
        ge=1,
        description="Page number",
    ),

    page_size: int = Query(
        default=20,
        ge=1,
        le=100,
        description="Number of records per page",
    ),

    sort_by: str = Query(
        default="assigned_date",
        description="Allowed: assigned_date, completed_date, decision_title",
    ),

    sort_order: str = Query(
        default="desc",
        description="Allowed: asc, desc",
    ),

    current_user: User = Depends(get_current_user),

    db: Session = Depends(get_db),
):
    # -----------------------------------------------------
    # VALIDATE APPROVAL STATUS
    # -----------------------------------------------------

    allowed_statuses = {
        "Pending",
        "Approved",
        "Rejected",
    }

    if (
        approval_status is not None
        and approval_status not in allowed_statuses
    ):
        raise HTTPException(
            status_code=422,
            detail=(
                "Invalid approval status. Allowed values: "
                "Pending, Approved, Rejected"
            ),
        )

    # -----------------------------------------------------
    # VALIDATE DATE RANGE
    # -----------------------------------------------------

    if start_date and end_date and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be greater than end_date",
        )

    # -----------------------------------------------------
    # VALIDATE SORTING
    # -----------------------------------------------------

    allowed_sort_fields = {
        "assigned_date": Approval.created_at,
        "completed_date": Approval.completed_at,
        "decision_title": Decision.title,
    }

    if sort_by not in allowed_sort_fields:
        raise HTTPException(
            status_code=422,
            detail=(
                "Invalid sorting field. Allowed values: "
                "assigned_date, completed_date, decision_title"
            ),
        )

    sort_order = sort_order.lower()

    if sort_order not in {"asc", "desc"}:
        raise HTTPException(
            status_code=422,
            detail="sort_order must be either 'asc' or 'desc'",
        )

    # -----------------------------------------------------
    # BASE QUERY
    # -----------------------------------------------------

    query = (
        db.query(Approval)
        .join(
            Decision,
            Approval.decision_id == Decision.id,
        )
        .join(
            User,
            Approval.reviewer_id == User.id,
        )
    )

    # -----------------------------------------------------
    # FILTERS
    # -----------------------------------------------------

    if approval_status:
        query = query.filter(
            Approval.status == approval_status
        )

    if reviewer:
        query = query.filter(
            Approval.reviewer_id == reviewer
        )

    if decision:
        query = query.filter(
            Approval.decision_id == decision
        )

    if start_date:
        query = query.filter(
            Approval.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            Approval.created_at <= end_date
        )

    # -----------------------------------------------------
    # TOTAL RECORDS
    # -----------------------------------------------------

    total_records = query.count()

    # -----------------------------------------------------
    # SUMMARY
    #
    # Summary uses all filtered records before pagination.
    # -----------------------------------------------------

    total_approvals = total_records

    pending_approvals = query.filter(
        Approval.status == "Pending"
    ).count()

    approved_approvals = query.filter(
        Approval.status == "Approved"
    ).count()

    rejected_approvals = query.filter(
        Approval.status == "Rejected"
    ).count()

    # -----------------------------------------------------
    # AVERAGE TURNAROUND TIME
    # -----------------------------------------------------

    completed_approvals = (
        query
        .filter(
            Approval.completed_at.isnot(None)
        )
        .all()
    )

    turnaround_values = []

    for approval in completed_approvals:
        if approval.completed_at and approval.created_at:
            duration = (
                approval.completed_at
                - approval.created_at
            )

            turnaround_values.append(
                duration.total_seconds() / 3600
            )

    if turnaround_values:
        average_turnaround = round(
            sum(turnaround_values)
            / len(turnaround_values),
            2,
        )
    else:
        average_turnaround = None

    # -----------------------------------------------------
    # APPROVAL COMPLETION RATE
    # -----------------------------------------------------

    if total_approvals > 0:
        completed_count = (
            approved_approvals
            + rejected_approvals
        )

        completion_rate = round(
            (completed_count / total_approvals) * 100,
            2,
        )
    else:
        completion_rate = 0.0

    # -----------------------------------------------------
    # SORT
    # -----------------------------------------------------

    sort_column = allowed_sort_fields[sort_by]

    if sort_order == "asc":
        query = query.order_by(
            sort_column.asc()
        )
    else:
        query = query.order_by(
            sort_column.desc()
        )

    # -----------------------------------------------------
    # PAGINATION
    # -----------------------------------------------------

    offset = (page - 1) * page_size

    approvals = (
        query
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # -----------------------------------------------------
    # BUILD REPORT DATA
    # -----------------------------------------------------

    report_data = []

    for approval in approvals:

        turnaround_time = None

        if (
            approval.completed_at
            and approval.created_at
        ):
            duration = (
                approval.completed_at
                - approval.created_at
            )

            turnaround_time = round(
                duration.total_seconds() / 3600,
                2,
            )

        report_data.append(
            ApprovalReportItem(
                approval_id=approval.id,

                decision_id=approval.decision_id,

                decision_title=(
                    approval.decision.title
                    if approval.decision
                    else ""
                ),

                reviewer=(
                    approval.reviewer.full_name
                    if approval.reviewer
                    else None
                ),

                approval_level=None,

                approval_status=approval.status,

                assigned_date=approval.created_at,

                completed_date=approval.completed_at,

                approval_turnaround_time_hours=(
                    turnaround_time
                ),
            )
        )

    # -----------------------------------------------------
    # TOTAL PAGES
    # -----------------------------------------------------

    total_pages = (
        ceil(total_records / page_size)
        if total_records > 0
        else 0
    )

    # -----------------------------------------------------
    # RESPONSE
    # -----------------------------------------------------

    return ApprovalReportResponse(
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages,
        sort_by=sort_by,
        sort_order=sort_order,

        summary=ApprovalReportSummary(
            total_approvals=total_approvals,
            pending_approvals=pending_approvals,
            approved_approvals=approved_approvals,
            rejected_approvals=rejected_approvals,
            average_approval_turnaround_time_hours=(
                average_turnaround
            ),
            approval_completion_rate=completion_rate,
        ),

        data=report_data,
    )
# ============================================================
# TEAM REPORTS
# TEAM = USER DEPARTMENT
# MANAGER / ADMINISTRATOR ONLY
# ============================================================

@router.get("/teams", response_model=TeamReportResponse)
def get_team_reports(
    team: str | None = Query(default=None),
    start_date: datetime | None = Query(default=None),
    end_date: datetime | None = Query(default=None),
    status: str | None = Query(default=None),
    category: str | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    sort_by: str = Query(default="team_name"),
    sort_order: str = Query(default="asc"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ========================================================
    # RBAC
    # ========================================================

    if current_user.role.lower() not in {
        "manager",
        "admin",
        "administrator",
    }:
        raise HTTPException(
            status_code=403,
            detail="Manager or Administrator access required",
        )

    # ========================================================
    # STATUS VALIDATION
    # ========================================================

    allowed_statuses = {
        "Draft",
        "Under Review",
        "Approved",
        "Rejected",
        "Archived",
    }

    if status is not None and status not in allowed_statuses:
        raise HTTPException(
            status_code=422,
            detail="Invalid status",
        )

    # ========================================================
    # DATE VALIDATION
    # ========================================================

    if start_date and end_date and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be greater than end_date",
        )

    # ========================================================
    # GET TEAMS FROM USER DEPARTMENTS
    # ========================================================

    team_query = (
        db.query(User.department)
        .filter(User.department.isnot(None))
        .distinct()
    )

    if team:
        team_query = team_query.filter(
            func.lower(User.department) == team.lower()
        )

    departments = [
        row[0]
        for row in team_query.all()
    ]

    # ========================================================
    # BUILD TEAM REPORT DATA
    # ========================================================

    report_data = []

    for department in departments:

        # ----------------------------------------------------
        # MEMBERS
        # ----------------------------------------------------

        member_count = (
            db.query(func.count(User.id))
            .filter(User.department == department)
            .scalar()
            or 0
        )

        # ----------------------------------------------------
        # DECISION QUERY
        # ----------------------------------------------------

        decision_query = (
            db.query(Decision)
            .join(
                User,
                Decision.created_by == User.id,
            )
            .filter(User.department == department)
        )

        if start_date:
            decision_query = decision_query.filter(
                Decision.created_at >= start_date
            )

        if end_date:
            decision_query = decision_query.filter(
                Decision.created_at <= end_date
            )

        if status:
            decision_query = decision_query.filter(
                Decision.status == status
            )

        if category:
            decision_query = decision_query.filter(
                Decision.category == category
            )

        decisions = decision_query.all()

        total_decisions = len(decisions)

        approved_decisions = sum(
            1 for d in decisions
            if d.status == "Approved"
        )

        rejected_decisions = sum(
            1 for d in decisions
            if d.status == "Rejected"
        )

        pending_decisions = sum(
            1 for d in decisions
            if d.status in {
                "Draft",
                "Under Review",
            }
        )

        # ----------------------------------------------------
        # APPROVALS
        # ----------------------------------------------------

        decision_ids = [d.id for d in decisions]

        if decision_ids:
            approvals = (
                db.query(Approval)
                .filter(
                    Approval.decision_id.in_(decision_ids)
                )
                .all()
            )
        else:
            approvals = []

        total_approvals = len(approvals)

        approved_approvals = sum(
            1 for a in approvals
            if a.status == "Approved"
        )

        rejected_approvals = sum(
            1 for a in approvals
            if a.status == "Rejected"
        )

        pending_approvals = sum(
            1 for a in approvals
            if a.status == "Pending"
        )

        # ----------------------------------------------------
        # APPROVAL COMPLETION RATE
        # ----------------------------------------------------

        if total_approvals > 0:
            approval_completion_rate = round(
                (
                    (approved_approvals + rejected_approvals)
                    / total_approvals
                ) * 100,
                2,
            )
        else:
            approval_completion_rate = 0.0

        # ----------------------------------------------------
        # AVERAGE APPROVAL TURNAROUND
        # ----------------------------------------------------

        completed_approvals = [
            a for a in approvals
            if a.completed_at is not None
        ]

        if completed_approvals:
            total_hours = sum(
                (
                    a.completed_at - a.created_at
                ).total_seconds() / 3600
                for a in completed_approvals
            )

            average_turnaround = round(
                total_hours / len(completed_approvals),
                2,
            )
        else:
            average_turnaround = None

        report_data.append(
            TeamReportItem(
                team_name=department,
                member_count=member_count,
                total_decisions=total_decisions,
                approved_decisions=approved_decisions,
                rejected_decisions=rejected_decisions,
                pending_decisions=pending_decisions,
                total_approvals=total_approvals,
                approved_approvals=approved_approvals,
                rejected_approvals=rejected_approvals,
                pending_approvals=pending_approvals,
                approval_completion_rate=approval_completion_rate,
                average_approval_turnaround_time_hours=average_turnaround,
            )
        )

    # ========================================================
    # CONTROLLED SORTING
    # ========================================================

    allowed_sort_fields = {
        "team_name": lambda x: x.team_name.lower(),
        "member_count": lambda x: x.member_count,
        "total_decisions": lambda x: x.total_decisions,
        "approved_decisions": lambda x: x.approved_decisions,
        "rejected_decisions": lambda x: x.rejected_decisions,
        "pending_decisions": lambda x: x.pending_decisions,
        "total_approvals": lambda x: x.total_approvals,
        "approval_completion_rate": lambda x: x.approval_completion_rate,
        "average_approval_turnaround_time_hours": (
            lambda x: (
                x.average_approval_turnaround_time_hours
                if x.average_approval_turnaround_time_hours is not None
                else 0
            )
        ),
    }

    if sort_by not in allowed_sort_fields:
        raise HTTPException(
            status_code=422,
            detail="Invalid sort_by",
        )

    if sort_order.lower() not in {"asc", "desc"}:
        raise HTTPException(
            status_code=422,
            detail="Invalid sort_order",
        )

    report_data.sort(
        key=allowed_sort_fields[sort_by],
        reverse=sort_order.lower() == "desc",
    )

    # ========================================================
    # PAGINATION
    # ========================================================

    total_records = len(report_data)

    offset = (page - 1) * page_size

    paginated_data = report_data[
        offset: offset + page_size
    ]

    total_pages = (
        (total_records + page_size - 1) // page_size
        if total_records > 0
        else 0
    )

    # ========================================================
    # RESPONSE
    # ========================================================

    return TeamReportResponse(
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages,
        sort_by=sort_by,
        sort_order=sort_order.lower(),
        data=paginated_data,
    )
# ============================================================
# AUDIT REPORTS
# ADMINISTRATOR ONLY
# ============================================================

@router.get("/audit", response_model=AuditReportResponse)
def get_audit_reports(
    user_id: int | None = Query(default=None, ge=1),
    action: str | None = Query(default=None),
    entity_type: str | None = Query(default=None),
    entity_id: int | None = Query(default=None, ge=1),
    start_date: datetime | None = Query(default=None),
    end_date: datetime | None = Query(default=None),
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=20, ge=1, le=100),
    sort_by: str = Query(default="timestamp"),
    sort_order: str = Query(default="desc"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):

    # ========================================================
    # ADMIN AUTHORIZATION
    # ========================================================

    if current_user.role.lower() not in {
        "admin",
        "administrator",
    }:
        raise HTTPException(
            status_code=403,
            detail="Administrator access required",
        )

    # ========================================================
    # VALIDATE ACTION
    # ========================================================

    if action is not None:
        action = action.upper()

        if action not in ALLOWED_ACTIONS:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid audit action: {action}",
            )

    # ========================================================
    # VALIDATE ENTITY TYPE
    # ========================================================

    if entity_type is not None:

        if entity_type not in ALLOWED_ENTITY_TYPES:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid entity type: {entity_type}",
            )

    # ========================================================
    # VALIDATE DATE RANGE
    # ========================================================

    if start_date and end_date and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be after end_date",
        )

    # ========================================================
    # CONTROLLED SORTING
    # ========================================================

    allowed_sort_fields = {
        "timestamp": AuditLog.created_at,
        "action": AuditLog.action,
        "entity_type": AuditLog.entity_type,
        "entity_id": AuditLog.entity_id,
        "user": AuditLog.user_id,
    }

    if sort_by not in allowed_sort_fields:
        raise HTTPException(
            status_code=422,
            detail="Invalid sort_by",
        )

    if sort_order.lower() not in {"asc", "desc"}:
        raise HTTPException(
            status_code=422,
            detail="Invalid sort_order",
        )

    # ========================================================
    # BASE QUERY
    # ========================================================

    query = (
        db.query(
            AuditLog,
            User.full_name,
        )
        .outerjoin(
            User,
            AuditLog.user_id == User.id,
        )
    )

    # ========================================================
    # FILTERS
    # ========================================================

    if user_id is not None:
        query = query.filter(
            AuditLog.user_id == user_id
        )

    if action is not None:
        query = query.filter(
            AuditLog.action == action
        )

    if entity_type is not None:
        query = query.filter(
            AuditLog.entity_type == entity_type
        )

    if entity_id is not None:
        query = query.filter(
            AuditLog.entity_id == entity_id
        )

    if start_date is not None:
        query = query.filter(
            AuditLog.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            AuditLog.created_at <= end_date
        )

    # ========================================================
    # TOTAL RECORDS
    # ========================================================

    total_records = query.count()

    # ========================================================
    # SORTING
    # ========================================================

    sort_column = allowed_sort_fields[sort_by]

    if sort_order.lower() == "desc":
        query = query.order_by(
            sort_column.desc()
        )
    else:
        query = query.order_by(
            sort_column.asc()
        )

    # ========================================================
    # PAGINATION
    # ========================================================

    offset = (page - 1) * page_size

    results = (
        query
        .offset(offset)
        .limit(page_size)
        .all()
    )

    # ========================================================
    # BUILD RESPONSE
    # ========================================================

    data = []

    for audit_log, user_name in results:

        data.append(
            AuditReportItem(
                user=user_name,
                action=audit_log.action,
                entity_type=audit_log.entity_type,
                entity_id=audit_log.entity_id,
                description=audit_log.description,
                timestamp=audit_log.created_at,
                ip_address=audit_log.ip_address,
            )
        )

    # ========================================================
    # TOTAL PAGES
    # ========================================================

    total_pages = (
        (total_records + page_size - 1) // page_size
        if total_records > 0
        else 0
    )

    # ========================================================
    # RESPONSE
    # ========================================================

    return AuditReportResponse(
        page=page,
        page_size=page_size,
        total_records=total_records,
        total_pages=total_pages,
        sort_by=sort_by,
        sort_order=sort_order.lower(),
        data=data,
    )
# ========================================================
# DECISION REPORT PDF EXPORT
# ========================================================

@router.get("/decisions/pdf")
def export_decision_report_pdf(
    category: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    created_by: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    tag: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "employee",
        "reviewer",
        "manager",
        "administrator",
        "admin",
    }:
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    valid_statuses = {
        "Draft",
        "Under Review",
        "Approved",
        "Rejected",
        "Archived",
    }

    if status is not None and status not in valid_statuses:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid status: {status}"
        )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # QUERY
    # ----------------------------------------------------

    query = (
        db.query(Decision, User.full_name)
        .join(User, Decision.created_by == User.id)
    )

    if category is not None:
        query = query.filter(Decision.category == category)

    if status is not None:
        query = query.filter(Decision.status == status)

    if created_by is not None:
        query = query.filter(Decision.created_by == created_by)

    if start_date is not None:
        query = query.filter(
            Decision.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            Decision.created_at <= end_date
        )

    if tag is not None:
        query = query.join(
            decision_tags,
            Decision.id == decision_tags.c.decision_id
        ).join(
            Tag,
            Tag.id == decision_tags.c.tag_id
        ).filter(
            func.lower(Tag.name) == tag.lower()
        )

    query = query.order_by(Decision.created_at.desc())

    results = query.all()

    # ----------------------------------------------------
    # CREATE PDF
    # ----------------------------------------------------

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    normal_style = styles["BodyText"]

    elements = []

    elements.append(
        Paragraph(
            "Decision Report",
            title_style
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"Total Decisions: {len(results)}",
            normal_style
        )
    )

    elements.append(Spacer(1, 10))

    # ----------------------------------------------------
    # SUMMARY
    # ----------------------------------------------------

    summary_data = [
        ["Status", "Count"],
        [
            "Draft",
            sum(
                1 for decision, _ in results
                if decision.status == "Draft"
            ),
        ],
        [
            "Under Review",
            sum(
                1 for decision, _ in results
                if decision.status == "Under Review"
            ),
        ],
        [
            "Approved",
            sum(
                1 for decision, _ in results
                if decision.status == "Approved"
            ),
        ],
        [
            "Rejected",
            sum(
                1 for decision, _ in results
                if decision.status == "Rejected"
            ),
        ],
        [
            "Archived",
            sum(
                1 for decision, _ in results
                if decision.status == "Archived"
            ),
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[120, 80]
    )

    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("TOPPADDING", (0, 0), (-1, 0), 6),
        ])
    )

    elements.append(summary_table)

    elements.append(Spacer(1, 15))

    # ----------------------------------------------------
    # DETAIL TABLE
    # ----------------------------------------------------

    table_data = [
        [
            "ID",
            "Title",
            "Category",
            "Status",
            "Created By",
            "Created Date",
            "Updated Date",
            "Alternatives",
            "Approvals",
            "Tags",
        ]
    ]

    for decision, creator_name in results:

        tag_names = ", ".join(
            tag.name for tag in decision.tags
        )

        row = [
            str(decision.id),
            Paragraph(
                decision.title or "",
                normal_style
            ),
            decision.category or "",
            decision.status or "",
            creator_name or "",
            decision.created_at.strftime(
                "%Y-%m-%d %H:%M"
            ),
            decision.updated_at.strftime(
                "%Y-%m-%d %H:%M"
            ),
            str(len(decision.alternatives)),
            str(len(decision.approvals)),
            Paragraph(
                tag_names,
                normal_style
            ),
        ]

        table_data.append(row)

    report_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            35,
            150,
            75,
            75,
            90,
            85,
            85,
            65,
            55,
            100,
        ],
    )

    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(report_table)

    # ----------------------------------------------------
    # BUILD PDF
    # ----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                "attachment; filename=decision_report.pdf"
            )
        },
    )
# ========================================================
# APPROVAL REPORT PDF EXPORT
# ========================================================

@router.get("/approvals/pdf")
def export_approval_report_pdf(
    approval_status: Optional[str] = Query(default=None),
    reviewer: Optional[int] = Query(default=None, ge=1),
    decision: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "employee",
        "reviewer",
        "manager",
        "administrator",
        "admin",
    }:
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    valid_statuses = {
        "Pending",
        "Approved",
        "Rejected",
    }

    if approval_status is not None:
        if approval_status not in valid_statuses:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid approval status: {approval_status}"
            )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # QUERY
    # ----------------------------------------------------

    query = (
        db.query(Approval, Decision, User.full_name)
        .join(
            Decision,
            Approval.decision_id == Decision.id
        )
        .join(
            User,
            Approval.reviewer_id == User.id
        )
    )

    if approval_status is not None:
        query = query.filter(
            Approval.status == approval_status
        )

    if reviewer is not None:
        query = query.filter(
            Approval.reviewer_id == reviewer
        )

    if decision is not None:
        query = query.filter(
            Approval.decision_id == decision
        )

    if start_date is not None:
        query = query.filter(
            Approval.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            Approval.created_at <= end_date
        )

    query = query.order_by(
        Approval.created_at.desc()
    )

    results = query.all()

    # ----------------------------------------------------
    # PDF
    # ----------------------------------------------------

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    normal_style = styles["BodyText"]

    elements = []

    elements.append(
        Paragraph(
            "Approval Report",
            title_style
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"Total Approvals: {len(results)}",
            normal_style
        )
    )

    elements.append(Spacer(1, 10))

    # ----------------------------------------------------
    # SUMMARY
    # ----------------------------------------------------

    pending_count = sum(
        1 for approval, _, _ in results
        if approval.status == "Pending"
    )

    approved_count = sum(
        1 for approval, _, _ in results
        if approval.status == "Approved"
    )

    rejected_count = sum(
        1 for approval, _, _ in results
        if approval.status == "Rejected"
    )

    completed = [
        approval
        for approval, _, _ in results
        if approval.completed_at is not None
    ]

    turnaround_values = []

    for approval in completed:
        turnaround = (
            approval.completed_at - approval.created_at
        ).total_seconds() / 3600

        turnaround_values.append(turnaround)

    average_turnaround = (
        sum(turnaround_values) / len(turnaround_values)
        if turnaround_values
        else 0
    )

    completion_rate = (
        (len(completed) / len(results)) * 100
        if results
        else 0
    )

    summary_data = [
        ["Metric", "Value"],
        ["Total Approvals", str(len(results))],
        ["Pending", str(pending_count)],
        ["Approved", str(approved_count)],
        ["Rejected", str(rejected_count)],
        [
            "Average Turnaround (Hours)",
            f"{average_turnaround:.2f}"
        ],
        [
            "Completion Rate",
            f"{completion_rate:.2f}%"
        ],
    ]

    summary_table = Table(
        summary_data,
        colWidths=[200, 100]
    )

    summary_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ])
    )

    elements.append(summary_table)

    elements.append(Spacer(1, 15))

    # ----------------------------------------------------
    # DETAIL TABLE
    # ----------------------------------------------------

    table_data = [
        [
            "Approval ID",
            "Decision ID",
            "Decision Title",
            "Reviewer",
            "Approval Level",
            "Status",
            "Assigned Date",
            "Completed Date",
            "Turnaround (Hours)",
        ]
    ]

    for approval, decision_obj, reviewer_name in results:

        turnaround = None

        if approval.completed_at is not None:
            turnaround = (
                approval.completed_at - approval.created_at
            ).total_seconds() / 3600

        table_data.append([
            str(approval.id),
            str(decision_obj.id),
            Paragraph(
                decision_obj.title or "",
                normal_style
            ),
            reviewer_name or "",
            "N/A",
            approval.status or "",
            approval.created_at.strftime(
                "%Y-%m-%d %H:%M"
            ),
            (
                approval.completed_at.strftime(
                    "%Y-%m-%d %H:%M"
                )
                if approval.completed_at
                else "N/A"
            ),
            (
                f"{turnaround:.2f}"
                if turnaround is not None
                else "N/A"
            ),
        ])

    report_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            55,
            55,
            140,
            90,
            70,
            65,
            90,
            90,
            80,
        ],
    )

    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(report_table)

    # ----------------------------------------------------
    # BUILD PDF
    # ----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                "attachment; filename=approval_report.pdf"
            )
        },
    )
# ========================================================
# TEAM REPORT PDF EXPORT
# ========================================================

@router.get("/teams/pdf")
def export_team_report_pdf(
    team: Optional[str] = Query(default=None),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    status: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "manager",
        "administrator",
        "admin",
    }:
        raise HTTPException(
            status_code=403,
            detail="Manager or Administrator access required"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    valid_statuses = {
        "Draft",
        "Under Review",
        "Approved",
        "Rejected",
        "Archived",
    }

    if status is not None and status not in valid_statuses:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid status: {status}"
        )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # FIND TEAMS
    # ----------------------------------------------------

    team_query = db.query(User.department).distinct()

    if team is not None:
        team_query = team_query.filter(
            User.department == team
        )

    teams = [
        row[0]
        for row in team_query.all()
        if row[0] is not None
    ]

    # ----------------------------------------------------
    # BUILD REPORT DATA
    # ----------------------------------------------------

    report_data = []

    for team_name in teams:

        # Team members
        members = db.query(User).filter(
            User.department == team_name
        ).all()

        member_ids = [member.id for member in members]

        # Decisions created by team members
        decision_query = db.query(Decision).filter(
            Decision.created_by.in_(member_ids)
        )

        if start_date is not None:
            decision_query = decision_query.filter(
                Decision.created_at >= start_date
            )

        if end_date is not None:
            decision_query = decision_query.filter(
                Decision.created_at <= end_date
            )

        if status is not None:
            decision_query = decision_query.filter(
                Decision.status == status
            )

        if category is not None:
            decision_query = decision_query.filter(
                Decision.category == category
            )

        decisions = decision_query.all()

        decision_ids = [decision.id for decision in decisions]

        # Decision statistics
        approved_decisions = sum(
            1 for decision in decisions
            if decision.status == "Approved"
        )

        rejected_decisions = sum(
            1 for decision in decisions
            if decision.status == "Rejected"
        )

        pending_decisions = sum(
            1 for decision in decisions
            if decision.status in {
                "Draft",
                "Under Review",
            }
        )

        # ------------------------------------------------
        # APPROVAL STATISTICS
        # ------------------------------------------------

        if decision_ids:
            approvals = db.query(Approval).filter(
                Approval.decision_id.in_(decision_ids)
            ).all()
        else:
            approvals = []

        approved_approvals = sum(
            1 for approval in approvals
            if approval.status == "Approved"
        )

        rejected_approvals = sum(
            1 for approval in approvals
            if approval.status == "Rejected"
        )

        pending_approvals = sum(
            1 for approval in approvals
            if approval.status == "Pending"
        )

        completed_approvals = [
            approval
            for approval in approvals
            if approval.completed_at is not None
        ]

        turnaround_values = []

        for approval in completed_approvals:
            turnaround = (
                approval.completed_at -
                approval.created_at
            ).total_seconds() / 3600

            turnaround_values.append(turnaround)

        average_turnaround = (
            sum(turnaround_values) /
            len(turnaround_values)
            if turnaround_values
            else None
        )

        completion_rate = (
            (
                len(completed_approvals) /
                len(approvals)
            ) * 100
            if approvals
            else 0
        )

        report_data.append({
            "team_name": team_name,
            "member_count": len(members),
            "total_decisions": len(decisions),
            "approved_decisions": approved_decisions,
            "rejected_decisions": rejected_decisions,
            "pending_decisions": pending_decisions,
            "total_approvals": len(approvals),
            "approved_approvals": approved_approvals,
            "rejected_approvals": rejected_approvals,
            "pending_approvals": pending_approvals,
            "completion_rate": completion_rate,
            "average_turnaround": average_turnaround,
        })

    # ----------------------------------------------------
    # CREATE PDF
    # ----------------------------------------------------

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    normal_style = styles["BodyText"]

    elements = []

    elements.append(
        Paragraph(
            "Team Report",
            title_style
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"Total Teams: {len(report_data)}",
            normal_style
        )
    )

    elements.append(Spacer(1, 15))

    # ----------------------------------------------------
    # TEAM TABLE
    # ----------------------------------------------------

    table_data = [
        [
            "Team",
            "Members",
            "Decisions",
            "Approved",
            "Rejected",
            "Pending",
            "Approvals",
            "Approved",
            "Rejected",
            "Pending",
            "Completion %",
            "Avg Turnaround",
        ]
    ]

    for item in report_data:

        avg_turnaround = (
            f"{item['average_turnaround']:.2f}"
            if item["average_turnaround"] is not None
            else "N/A"
        )

        table_data.append([
            item["team_name"],
            str(item["member_count"]),
            str(item["total_decisions"]),
            str(item["approved_decisions"]),
            str(item["rejected_decisions"]),
            str(item["pending_decisions"]),
            str(item["total_approvals"]),
            str(item["approved_approvals"]),
            str(item["rejected_approvals"]),
            str(item["pending_approvals"]),
            f"{item['completion_rate']:.2f}%",
            avg_turnaround,
        ])

    report_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            80,
            50,
            60,
            55,
            55,
            55,
            55,
            55,
            55,
            55,
            65,
            75,
        ],
    )

    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("ALIGN", (1, 1), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(report_table)

    # ----------------------------------------------------
    # BUILD PDF
    # ----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                "attachment; filename=team_report.pdf"
            )
        },
    )
# ========================================================
# AUDIT REPORT PDF EXPORT
# ========================================================

@router.get("/audit/pdf")
def export_audit_report_pdf(
    user_id: Optional[int] = Query(default=None, ge=1),
    action: Optional[str] = Query(default=None),
    entity_type: Optional[str] = Query(default=None),
    entity_id: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "admin",
        "administrator",
    }:
        raise HTTPException(
            status_code=403,
            detail="Administrator access required"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    if action is not None:
        action = action.upper()

        if action not in ALLOWED_ACTIONS:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid audit action: {action}"
            )

    if entity_type is not None:
        if entity_type not in ALLOWED_ENTITY_TYPES:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid entity type: {entity_type}"
            )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # QUERY
    # ----------------------------------------------------

    query = (
        db.query(AuditLog, User.full_name)
        .outerjoin(
            User,
            AuditLog.user_id == User.id
        )
    )

    if user_id is not None:
        query = query.filter(
            AuditLog.user_id == user_id
        )

    if action is not None:
        query = query.filter(
            AuditLog.action == action
        )

    if entity_type is not None:
        query = query.filter(
            AuditLog.entity_type == entity_type
        )

    if entity_id is not None:
        query = query.filter(
            AuditLog.entity_id == entity_id
        )

    if start_date is not None:
        query = query.filter(
            AuditLog.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            AuditLog.created_at <= end_date
        )

    query = query.order_by(
        AuditLog.created_at.desc()
    )

    results = query.all()

    # ----------------------------------------------------
    # CREATE PDF
    # ----------------------------------------------------

    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=25,
        leftMargin=25,
        topMargin=25,
        bottomMargin=25,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]
    normal_style = styles["BodyText"]

    elements = []

    elements.append(
        Paragraph(
            "Audit Report",
            title_style
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"Total Audit Records: {len(results)}",
            normal_style
        )
    )

    elements.append(Spacer(1, 15))

    # ----------------------------------------------------
    # AUDIT TABLE
    # ----------------------------------------------------

    table_data = [
        [
            "User",
            "Action",
            "Entity Type",
            "Entity ID",
            "Description",
            "Timestamp",
            "IP Address",
        ]
    ]

    for audit_log, user_name in results:

        table_data.append([
            user_name or "System",
            audit_log.action or "",
            audit_log.entity_type or "",
            (
                str(audit_log.entity_id)
                if audit_log.entity_id is not None
                else "N/A"
            ),
            Paragraph(
                audit_log.description or "",
                normal_style
            ),
            audit_log.created_at.strftime(
                "%Y-%m-%d %H:%M:%S"
            ),
            audit_log.ip_address or "N/A",
        ])

    report_table = Table(
        table_data,
        repeatRows=1,
        colWidths=[
            90,
            65,
            85,
            55,
            220,
            110,
            90,
        ],
    )

    report_table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.black),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ])
    )

    elements.append(report_table)

    # ----------------------------------------------------
    # BUILD PDF
    # ----------------------------------------------------

    document.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": (
                "attachment; filename=audit_report.pdf"
            )
        },
    )
# ========================================================
# DECISION REPORT EXCEL EXPORT
# ========================================================

@router.get("/decisions/excel")
def export_decision_report_excel(
    category: Optional[str] = Query(default=None),
    status: Optional[str] = Query(default=None),
    created_by: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    tag: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "employee",
        "reviewer",
        "manager",
        "administrator",
        "admin",
    }:
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    valid_statuses = {
        "Draft",
        "Under Review",
        "Approved",
        "Rejected",
        "Archived",
    }

    if status is not None and status not in valid_statuses:
        raise HTTPException(
            status_code=422,
            detail=f"Invalid status: {status}"
        )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # QUERY
    # ----------------------------------------------------

    query = (
        db.query(Decision, User.full_name)
        .join(User, Decision.created_by == User.id)
    )

    if category is not None:
        query = query.filter(
            Decision.category == category
        )

    if status is not None:
        query = query.filter(
            Decision.status == status
        )

    if created_by is not None:
        query = query.filter(
            Decision.created_by == created_by
        )

    if start_date is not None:
        query = query.filter(
            Decision.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            Decision.created_at <= end_date
        )

    if tag is not None:
        query = query.join(
            decision_tags,
            Decision.id == decision_tags.c.decision_id
        ).join(
            Tag,
            Tag.id == decision_tags.c.tag_id
        ).filter(
            func.lower(Tag.name) == tag.lower()
        )

    query = query.order_by(
        Decision.created_at.desc()
    )

    results = query.all()

    # ----------------------------------------------------
    # CREATE WORKBOOK
    # ----------------------------------------------------

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Decision Report"

    # ----------------------------------------------------
    # TITLE
    # ----------------------------------------------------

    worksheet["A1"] = "Decision Report"
    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A3"] = "Total Decisions"
    worksheet["B3"] = len(results)

    # ----------------------------------------------------
    # SUMMARY
    # ----------------------------------------------------

    worksheet["A5"] = "Status Summary"
    worksheet["A5"].font = Font(
        bold=True,
        size=13
    )

    summary_headers = [
        "Status",
        "Count"
    ]

    worksheet.append(summary_headers)

    for cell in worksheet[6]:
        cell.font = Font(bold=True)

    status_counts = {
        "Draft": 0,
        "Under Review": 0,
        "Approved": 0,
        "Rejected": 0,
        "Archived": 0,
    }

    for decision, _ in results:
        if decision.status in status_counts:
            status_counts[decision.status] += 1

    for status_name, count in status_counts.items():
        worksheet.append([
            status_name,
            count
        ])

    # ----------------------------------------------------
    # DETAIL TABLE
    # ----------------------------------------------------

    detail_start_row = worksheet.max_row + 2

    headers = [
        "Decision ID",
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date",
        "Updated Date",
        "Alternatives",
        "Approvals",
        "Tags",
    ]

    for column_index, header in enumerate(
        headers,
        start=1
    ):
        cell = worksheet.cell(
            row=detail_start_row,
            column=column_index
        )
        cell.value = header
        cell.font = Font(bold=True)

    # ----------------------------------------------------
    # DATA
    # ----------------------------------------------------

    for row_index, (decision, creator_name) in enumerate(
        results,
        start=detail_start_row + 1
    ):

        tag_names = ", ".join(
            tag.name
            for tag in decision.tags
        )

        values = [
            decision.id,
            decision.title,
            decision.category,
            decision.status,
            creator_name,
            decision.created_at,
            decision.updated_at,
            len(decision.alternatives),
            len(decision.approvals),
            tag_names,
        ]

        for column_index, value in enumerate(
            values,
            start=1
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value
            )

    # ----------------------------------------------------
    # FORMATTING
    # ----------------------------------------------------

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    column_widths = {
        "A": 14,
        "B": 30,
        "C": 18,
        "D": 18,
        "E": 25,
        "F": 22,
        "G": 22,
        "H": 15,
        "I": 15,
        "J": 30,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[
            column
        ].width = width

    # ----------------------------------------------------
    # SAVE TO MEMORY
    # ----------------------------------------------------

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    # ----------------------------------------------------
    # RESPONSE
    # ----------------------------------------------------

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename=decision_report.xlsx"
            )
        },
    )
# ========================================================
# APPROVAL REPORT EXCEL EXPORT
# ========================================================

@router.get("/approvals/excel")
def export_approval_report_excel(
    approval_status: Optional[str] = Query(default=None),
    reviewer: Optional[int] = Query(default=None, ge=1),
    decision: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = Query(default=None),
    end_date: Optional[datetime] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # ----------------------------------------------------
    # AUTHORIZATION
    # ----------------------------------------------------

    if current_user.role.lower() not in {
        "employee",
        "reviewer",
        "manager",
        "administrator",
        "admin",
    }:
        raise HTTPException(
            status_code=403,
            detail="Access denied"
        )

    # ----------------------------------------------------
    # VALIDATION
    # ----------------------------------------------------

    valid_statuses = {
        "Pending",
        "Approved",
        "Rejected",
    }

    if approval_status is not None:
        if approval_status not in valid_statuses:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid approval status: {approval_status}"
            )

    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # ----------------------------------------------------
    # QUERY
    # ----------------------------------------------------

    query = (
        db.query(Approval, Decision, User.full_name)
        .join(
            Decision,
            Approval.decision_id == Decision.id
        )
        .join(
            User,
            Approval.reviewer_id == User.id
        )
    )

    if approval_status is not None:
        query = query.filter(
            Approval.status == approval_status
        )

    if reviewer is not None:
        query = query.filter(
            Approval.reviewer_id == reviewer
        )

    if decision is not None:
        query = query.filter(
            Approval.decision_id == decision
        )

    if start_date is not None:
        query = query.filter(
            Approval.created_at >= start_date
        )

    if end_date is not None:
        query = query.filter(
            Approval.created_at <= end_date
        )

    query = query.order_by(
        Approval.created_at.desc()
    )

    results = query.all()

    # ----------------------------------------------------
    # CREATE WORKBOOK
    # ----------------------------------------------------

    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Approval Report"

    # ----------------------------------------------------
    # TITLE
    # ----------------------------------------------------

    worksheet["A1"] = "Approval Report"
    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet["A3"] = "Total Approvals"
    worksheet["B3"] = len(results)

    # ----------------------------------------------------
    # SUMMARY
    # ----------------------------------------------------

    worksheet["A5"] = "Approval Summary"
    worksheet["A5"].font = Font(
        bold=True,
        size=13
    )

    summary_headers = [
        "Metric",
        "Value"
    ]

    for column_index, header in enumerate(
        summary_headers,
        start=1
    ):
        cell = worksheet.cell(
            row=6,
            column=column_index
        )
        cell.value = header
        cell.font = Font(bold=True)

    pending_count = sum(
        1
        for approval, _, _ in results
        if approval.status == "Pending"
    )

    approved_count = sum(
        1
        for approval, _, _ in results
        if approval.status == "Approved"
    )

    rejected_count = sum(
        1
        for approval, _, _ in results
        if approval.status == "Rejected"
    )

    completed = [
        approval
        for approval, _, _ in results
        if approval.completed_at is not None
    ]

    turnaround_values = []

    for approval in completed:
        turnaround = (
            approval.completed_at -
            approval.created_at
        ).total_seconds() / 3600

        turnaround_values.append(turnaround)

    average_turnaround = (
        sum(turnaround_values) /
        len(turnaround_values)
        if turnaround_values
        else 0
    )

    completion_rate = (
        (
            len(completed) /
            len(results)
        ) * 100
        if results
        else 0
    )

    summary_rows = [
        ["Pending", pending_count],
        ["Approved", approved_count],
        ["Rejected", rejected_count],
        [
            "Average Turnaround (Hours)",
            round(average_turnaround, 2)
        ],
        [
            "Completion Rate (%)",
            round(completion_rate, 2)
        ],
    ]

    for row in summary_rows:
        worksheet.append(row)

    # ----------------------------------------------------
    # DETAIL TABLE
    # ----------------------------------------------------

    detail_start_row = worksheet.max_row + 2

    headers = [
        "Approval ID",
        "Decision ID",
        "Decision Title",
        "Reviewer",
        "Approval Level",
        "Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround (Hours)",
    ]

    for column_index, header in enumerate(
        headers,
        start=1
    ):
        cell = worksheet.cell(
            row=detail_start_row,
            column=column_index
        )
        cell.value = header
        cell.font = Font(bold=True)

    # ----------------------------------------------------
    # DATA
    # ----------------------------------------------------

    for row_index, (
        approval,
        decision_obj,
        reviewer_name
    ) in enumerate(
        results,
        start=detail_start_row + 1
    ):

        turnaround = None

        if approval.completed_at is not None:
            turnaround = (
                approval.completed_at -
                approval.created_at
            ).total_seconds() / 3600

        values = [
            approval.id,
            decision_obj.id,
            decision_obj.title,
            reviewer_name,
            "N/A",
            approval.status,
            approval.created_at,
            (
                approval.completed_at
                if approval.completed_at
                else None
            ),
            (
                round(turnaround, 2)
                if turnaround is not None
                else None
            ),
        ]

        for column_index, value in enumerate(
            values,
            start=1
        ):
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=value
            )

    # ----------------------------------------------------
    # FORMATTING
    # ----------------------------------------------------

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

    column_widths = {
        "A": 15,
        "B": 14,
        "C": 30,
        "D": 25,
        "E": 18,
        "F": 15,
        "G": 22,
        "H": 22,
        "I": 22,
    }

    for column, width in column_widths.items():
        worksheet.column_dimensions[
            column
        ].width = width

    # ----------------------------------------------------
    # SAVE
    # ----------------------------------------------------

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    # ----------------------------------------------------
    # RESPONSE
    # ----------------------------------------------------

    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": (
                "attachment; filename=approval_report.xlsx"
            )
        },
    )

@router.get("/teams/excel")
def export_team_report_excel(
    team: Optional[str] = Query(default=None),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    status: Optional[str] = Query(default=None),
    category: Optional[str] = Query(default=None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Authorization
    if current_user.role.lower() not in {"manager", "admin", "administrator"}:
        raise HTTPException(
            status_code=403,
            detail="Manager or Administrator access required"
        )

    # Date validation
    if start_date is not None and end_date is not None and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be after end_date"
        )

    # Get departments (used as teams)
    team_query = db.query(User.department).distinct()

    if team:
        team_query = team_query.filter(User.department == team)

    teams = [row[0] for row in team_query.all()]

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Team Report"

    # Title
    worksheet["A1"] = "Team Report"
    worksheet["A1"].font = Font(bold=True, size=16)

    worksheet["A2"] = "Generated By"
    worksheet["B2"] = current_user.full_name

    row = 4

    # Headers
    headers = [
        "Team Name",
        "Member Count",
        "Total Decisions",
        "Approved Decisions",
        "Rejected Decisions",
        "Pending Decisions",
        "Total Approvals",
        "Approved Approvals",
        "Rejected Approvals",
        "Pending Approvals",
        "Approval Completion Rate",
        "Average Approval Turnaround (Hours)",
    ]

    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    row += 1

    for team_name in teams:

        members = db.query(User.id).filter(
            User.department == team_name
        ).all()

        member_ids = [member[0] for member in members]

        if not member_ids:
            continue

        decision_query = db.query(Decision).filter(
            Decision.created_by.in_(member_ids)
        )

        if start_date is not None:
            decision_query = decision_query.filter(
                Decision.created_at >= start_date
            )

        if end_date is not None:
            decision_query = decision_query.filter(
                Decision.created_at <= end_date
            )

        if status is not None:
            decision_query = decision_query.filter(
                Decision.status == status
            )

        if category is not None:
            decision_query = decision_query.filter(
                Decision.category == category
            )

        decisions = decision_query.all()

        decision_ids = [decision.id for decision in decisions]

        total_decisions = len(decisions)
        approved_decisions = sum(
            1 for d in decisions if d.status == "Approved"
        )
        rejected_decisions = sum(
            1 for d in decisions if d.status == "Rejected"
        )
        pending_decisions = sum(
            1 for d in decisions
            if d.status in {"Draft", "Under Review"}
        )

        approvals = []

        if decision_ids:
            approvals = db.query(Approval).filter(
                Approval.decision_id.in_(decision_ids)
            ).all()

        total_approvals = len(approvals)
        approved_approvals = sum(
            1 for a in approvals if a.status == "Approved"
        )
        rejected_approvals = sum(
            1 for a in approvals if a.status == "Rejected"
        )
        pending_approvals = sum(
            1 for a in approvals if a.status == "Pending"
        )

        completed_approvals = [
            a for a in approvals
            if a.completed_at is not None
        ]

        completion_rate = (
            (len(completed_approvals) / total_approvals) * 100
            if total_approvals
            else 0
        )

        turnaround_values = []

        for approval in completed_approvals:
            turnaround = (
                approval.completed_at - approval.created_at
            ).total_seconds() / 3600

            turnaround_values.append(turnaround)

        average_turnaround = (
            sum(turnaround_values) / len(turnaround_values)
            if turnaround_values
            else None
        )

        values = [
            team_name,
            len(member_ids),
            total_decisions,
            approved_decisions,
            rejected_decisions,
            pending_decisions,
            total_approvals,
            approved_approvals,
            rejected_approvals,
            pending_approvals,
            round(completion_rate, 2),
            round(average_turnaround, 2)
            if average_turnaround is not None
            else None,
        ]

        for col, value in enumerate(values, start=1):
            worksheet.cell(row=row, column=col, value=value)

        row += 1

    # Formatting
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            35
        )

    worksheet.freeze_panes = "A5"

    # Create Excel response
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition": "attachment; filename=team_report.xlsx"
        },
    )
@router.get("/audit/excel")
def export_audit_report_excel(
    user_id: Optional[int] = Query(default=None, ge=1),
    action: Optional[str] = Query(default=None),
    entity_type: Optional[str] = Query(default=None),
    entity_id: Optional[int] = Query(default=None, ge=1),
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Administrator authorization
    if current_user.role.lower() not in {"admin", "administrator"}:
        raise HTTPException(
            status_code=403,
            detail="Administrator access required"
        )

    # Validate action
    if action is not None:
        action = action.upper()

        if action not in ALLOWED_ACTIONS:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid audit action: {action}"
            )

    # Validate entity type
    if entity_type is not None:
        if entity_type not in ALLOWED_ENTITY_TYPES:
            raise HTTPException(
                status_code=422,
                detail=f"Invalid entity type: {entity_type}"
            )

    # Validate date range
    if start_date is not None and end_date is not None:
        if start_date > end_date:
            raise HTTPException(
                status_code=422,
                detail="start_date cannot be after end_date"
            )

    # Build query
    query = (
        db.query(AuditLog, User)
        .outerjoin(User, AuditLog.user_id == User.id)
    )

    if user_id is not None:
        query = query.filter(AuditLog.user_id == user_id)

    if action is not None:
        query = query.filter(AuditLog.action == action)

    if entity_type is not None:
        query = query.filter(AuditLog.entity_type == entity_type)

    if entity_id is not None:
        query = query.filter(AuditLog.entity_id == entity_id)

    if start_date is not None:
        query = query.filter(AuditLog.created_at >= start_date)

    if end_date is not None:
        query = query.filter(AuditLog.created_at <= end_date)

    records = query.order_by(
        AuditLog.created_at.desc()
    ).all()

    # Create workbook
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Audit Report"

    # Title
    worksheet["A1"] = "Audit Report"
    worksheet["A1"].font = Font(bold=True, size=16)

    worksheet["A2"] = "Generated By"
    worksheet["B2"] = current_user.full_name

    worksheet["A3"] = "Total Records"
    worksheet["B3"] = len(records)

    # Headers
    headers = [
        "User",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "Timestamp",
        "IP Address",
    ]

    header_row = 5

    for col, header in enumerate(headers, start=1):
        cell = worksheet.cell(
            row=header_row,
            column=col,
            value=header
        )
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data
    row = header_row + 1

    for audit_log, user in records:
        values = [
            user.full_name if user else None,
            audit_log.action,
            audit_log.entity_type,
            audit_log.entity_id,
            audit_log.description,
            audit_log.created_at,
            audit_log.ip_address,
        ]

        for col, value in enumerate(values, start=1):
            worksheet.cell(
                row=row,
                column=col,
                value=value
            )

        row += 1

    # Format timestamp column
    for cell in worksheet["F"][5:]:
        if cell.value is not None:
            cell.number_format = "yyyy-mm-dd hh:mm:ss"

    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[column_letter].width = min(
            max_length + 2,
            50
        )

    # Freeze header
    worksheet.freeze_panes = "A6"

    # Create response
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
                "attachment; filename=audit_report.xlsx"
        },
    )