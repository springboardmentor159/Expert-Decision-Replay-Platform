from fastapi import APIRouter, Depends, Query, HTTPException
from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from sqlalchemy import func, or_, text

from typing import Optional
from datetime import date, datetime
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

from openpyxl import Workbook

from app.database import get_db
from app.models.decision import Decision
from app.models.approval import Approval
from app.models.user import User
from app.models.audit_log import AuditLog
from app.routers.auth import get_current_user


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


# ============================================================
# COMMON HELPERS
# ============================================================

VALID_DECISION_STATUSES = [
    "Draft",
    "Under Review",
    "Approved",
    "Rejected",
    "Archived"
]

VALID_SORT_FIELDS = [
    "created_at",
    "updated_at",
    "title"
]

VALID_APPROVAL_SORT_FIELDS = [
    "approval_date",
    "assigned_at",
    "completed_at"
]

VALID_TEAM_SORT_FIELDS = [
    "team_name"
]


def validate_date_range(
    start_date: Optional[date],
    end_date: Optional[date]
):
    if start_date and end_date and start_date > end_date:
        raise HTTPException(
            status_code=422,
            detail="start_date cannot be greater than end_date"
        )


def validate_status(status: Optional[str]):
    if status and status not in VALID_DECISION_STATUSES:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "Invalid status",
                "allowed_values": VALID_DECISION_STATUSES
            }
        )


def validate_pagination(
    page: int,
    page_size: int
):
    if page < 1:
        raise HTTPException(
            status_code=422,
            detail="page must be greater than or equal to 1"
        )

    if page_size < 1 or page_size > 100:
        raise HTTPException(
            status_code=422,
            detail="page_size must be between 1 and 100"
        )


def date_end_datetime(end_date: Optional[date]):
    """
    Converts an end date into the beginning of the next day so
    the complete end_date is included.
    """
    if not end_date:
        return None

    from datetime import timedelta

    return datetime.combine(
        end_date + timedelta(days=1),
        datetime.min.time()
    )


def check_manager_or_admin(current_user: User, message: str):
    if current_user.role not in [
        "Manager",
        "Administrator"
    ]:
        raise HTTPException(
            status_code=403,
            detail=message
        )


def get_decision_tags(db: Session, decision_id: int):
    """
    Reads tags using the existing decision_tags/tags tables.

    Expected structure:
        decision_tags.decision_id
        decision_tags.tag_id
        tags.id
        tags.name
    """

    try:
        result = db.execute(
            text("""
                SELECT t.name
                FROM tags t
                INNER JOIN decision_tags dt
                    ON dt.tag_id = t.id
                WHERE dt.decision_id = :decision_id
                ORDER BY t.name
            """),
            {"decision_id": decision_id}
        )

        return [
            row[0]
            for row in result.fetchall()
        ]

    except Exception:
        return []


def get_tag_filter_ids(
    db: Session,
    tag: Optional[str]
):
    if not tag:
        return None

    result = db.execute(
        text("""
            SELECT id
            FROM tags
            WHERE LOWER(name) = LOWER(:tag)
        """),
        {"tag": tag}
    )

    return [
        row[0]
        for row in result.fetchall()
    ]


def create_pdf(
    title,
    headers,
    rows,
    summary=None,
    filters=None
):
    buffer = BytesIO()

    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=20,
        leftMargin=20,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()

    elements = []

    elements.append(
        Paragraph(
            title,
            styles["Title"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    elements.append(
        Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            styles["Normal"]
        )
    )

    elements.append(
        Spacer(1, 10)
    )

    if filters:
        filter_text = " | ".join(
            f"{key}: {value}"
            for key, value in filters.items()
            if value is not None
        )

        if filter_text:
            elements.append(
                Paragraph(
                    f"Filters: {filter_text}",
                    styles["Normal"]
                )
            )

            elements.append(
                Spacer(1, 10)
            )

    if summary:
        summary_data = [
            ["Summary", "Value"]
        ]

        for key, value in summary.items():
            summary_data.append(
                [str(key), str(value)]
            )

        summary_table = Table(
            summary_data,
            repeatRows=1
        )

        summary_table.setStyle(
            TableStyle([
                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.grey
                ),
                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),
                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.black
                ),
                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),
                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                )
            ])
        )

        elements.append(summary_table)
        elements.append(Spacer(1, 15))

    table_data = [headers]

    for row in rows:
        table_data.append(
            [
                str(value) if value is not None else ""
                for value in row
            ]
        )

    if len(table_data) == 1:
        table_data.append(
            ["No records found"]
            + [""] * (len(headers) - 1)
        )

    table = Table(
        table_data,
        repeatRows=1
    )

    table.setStyle(
        TableStyle([
            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.darkgrey
            ),
            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),
            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),
            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.black
            ),
            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "TOP"
            ),
        ])
    )

    elements.append(table)

    document.build(elements)

    buffer.seek(0)

    return buffer


def create_excel(
    title,
    headers,
    rows,
    summary=None,
    filters=None
):
    workbook = Workbook()

    worksheet = workbook.active
    worksheet.title = "Report"

    worksheet["A1"] = title
    worksheet["A2"] = (
        f"Generated: "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )

    current_row = 4

    if filters:
        worksheet.cell(
            row=current_row,
            column=1,
            value="Filters"
        )

        current_row += 1

        for key, value in filters.items():
            if value is not None:
                worksheet.cell(
                    row=current_row,
                    column=1,
                    value=str(key)
                )

                worksheet.cell(
                    row=current_row,
                    column=2,
                    value=str(value)
                )

                current_row += 1

        current_row += 1

    if summary:
        worksheet.cell(
            row=current_row,
            column=1,
            value="Summary"
        )

        current_row += 1

        worksheet.cell(
            row=current_row,
            column=1,
            value="Metric"
        )

        worksheet.cell(
            row=current_row,
            column=2,
            value="Value"
        )

        current_row += 1

        for key, value in summary.items():
            worksheet.cell(
                row=current_row,
                column=1,
                value=str(key)
            )

            worksheet.cell(
                row=current_row,
                column=2,
                value=value
            )

            current_row += 1

        current_row += 1

    for column_number, header in enumerate(
        headers,
        start=1
    ):
        worksheet.cell(
            row=current_row,
            column=column_number,
            value=header
        )

    current_row += 1

    for row in rows:
        for column_number, value in enumerate(
            row,
            start=1
        ):
            worksheet.cell(
                row=current_row,
                column=column_number,
                value=value
            )

        current_row += 1

    for column in worksheet.columns:
        max_length = 0

        column_letter = column[0].column_letter

        for cell in column:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )

        worksheet.column_dimensions[
            column_letter
        ].width = min(max_length + 2, 40)

    buffer = BytesIO()

    workbook.save(buffer)

    buffer.seek(0)

    return buffer


# ============================================================
# TASK 1 - DECISION REPORT
# ============================================================

@router.get("/decisions")
def get_decision_report(
    category: Optional[str] = None,
    status: Optional[str] = None,
    created_by: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    tag: Optional[str] = None,

    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),

    sort_by: str = Query(
        "created_at",
        description=(
            "Allowed values: "
            "created_at, updated_at, title"
        )
    ),
    sort_order: str = Query(
        "desc",
        description="Allowed values: asc, desc"
    ),

    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    validate_date_range(
        start_date,
        end_date
    )

    validate_status(status)

    if sort_by not in VALID_SORT_FIELDS:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "Invalid sort field",
                "allowed_values": VALID_SORT_FIELDS
            }
        )

    if sort_order not in ["asc", "desc"]:
        raise HTTPException(
            status_code=422,
            detail="sort_order must be asc or desc"
        )

    query = db.query(Decision)

    if category:
        query = query.filter(
            Decision.category == category
        )

    if status:
        query = query.filter(
            Decision.status == status
        )

    if created_by:
        query = query.filter(
            Decision.created_by == created_by
        )

    if start_date:
        query = query.filter(
            Decision.created_at >=
            datetime.combine(
                start_date,
                datetime.min.time()
            )
        )

    end_datetime = date_end_datetime(end_date)

    if end_datetime:
        query = query.filter(
            Decision.created_at < end_datetime
        )

    if tag:
        tag_ids = get_tag_filter_ids(
            db,
            tag
        )

        if not tag_ids:
            return {
                "page": page,
                "page_size": page_size,
                "total_records": 0,
                "summary": {
                    "total": 0,
                    "draft": 0,
                    "under_review": 0,
                    "approved": 0,
                    "rejected": 0,
                    "archived": 0
                },
                "data": []
            }

        matching_decisions = db.execute(
            text("""
                SELECT DISTINCT decision_id
                FROM decision_tags
                WHERE tag_id IN :tag_ids
            """).bindparams(
                tag_ids=tuple(tag_ids)
            )
        )

        decision_ids = [
            row[0]
            for row in matching_decisions.fetchall()
        ]

        if not decision_ids:
            return {
                "page": page,
                "page_size": page_size,
                "total_records": 0,
                "summary": {
                    "total": 0,
                    "draft": 0,
                    "under_review": 0,
                    "approved": 0,
                    "rejected": 0,
                    "archived": 0
                },
                "data": []
            }

        query = query.filter(
            Decision.id.in_(decision_ids)
        )

    sort_column = getattr(
        Decision,
        sort_by
    )

    if sort_order == "asc":
        query = query.order_by(
            sort_column.asc()
        )
    else:
        query = query.order_by(
            sort_column.desc()
        )

    total_records = query.count()

    decisions = query.offset(
        (page - 1) * page_size
    ).limit(
        page_size
    ).all()

    # Summary respects the selected filters.
    summary_query = db.query(Decision)

    if category:
        summary_query = summary_query.filter(
            Decision.category == category
        )

    if status:
        summary_query = summary_query.filter(
            Decision.status == status
        )

    if created_by:
        summary_query = summary_query.filter(
            Decision.created_by == created_by
        )

    if start_date:
        summary_query = summary_query.filter(
            Decision.created_at >=
            datetime.combine(
                start_date,
                datetime.min.time()
            )
        )

    if end_datetime:
        summary_query = summary_query.filter(
            Decision.created_at < end_datetime
        )

    if tag:
        summary_query = summary_query.filter(
            Decision.id.in_(
                decision_ids
            )
        )

    summary_rows = summary_query.all()

    summary = {
        "total": len(summary_rows),
        "draft": sum(
            1 for d in summary_rows
            if d.status == "Draft"
        ),
        "under_review": sum(
            1 for d in summary_rows
            if d.status == "Under Review"
        ),
        "approved": sum(
            1 for d in summary_rows
            if d.status == "Approved"
        ),
        "rejected": sum(
            1 for d in summary_rows
            if d.status == "Rejected"
        ),
        "archived": sum(
            1 for d in summary_rows
            if d.status == "Archived"
        )
    }

    data = []

    for decision in decisions:

        creator = db.query(User).filter(
            User.id == decision.created_by
        ).first()

        alternative_count = 0

        if hasattr(decision, "alternatives"):
            alternative_count = len(
                decision.alternatives
            )

        approval_count = db.query(
            Approval
        ).filter(
            Approval.decision_id == decision.id
        ).count()

        tags = get_decision_tags(
            db,
            decision.id
        )

        data.append({
            "decision_id": decision.id,
            "title": decision.title,
            "category": decision.category,
            "status": decision.status,
            "created_by": (
                creator.full_name
                if creator
                else decision.created_by
            ),
            "created_date": decision.created_at,
            "updated_date": decision.updated_at,
            "number_alternatives": alternative_count,
            "number_approvals": approval_count,
            "tags": tags
        })

    return {
        "page": page,
        "page_size": page_size,
        "total_records": total_records,
        "summary": summary,
        "data": data
    }


# ============================================================
# TASK 2 - APPROVAL REPORT
# ============================================================

@router.get("/approvals")
def get_approval_report(
    status: Optional[str] = None,
    reviewer_id: Optional[int] = None,
    decision_id: Optional[int] = None,
    approval_level: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,

    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),

    sort_by: str = Query(
        "assigned_at",
        description=(
            "Allowed values: "
            "approval_date, assigned_at, completed_at"
        )
    ),

    sort_order: str = Query(
        "desc",
        description="Allowed values: asc, desc"
    ),

    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    validate_date_range(
        start_date,
        end_date
    )

    validate_pagination(
        page,
        page_size
    )

    if sort_by not in VALID_APPROVAL_SORT_FIELDS:
        raise HTTPException(
            status_code=422,
            detail={
                "error": "Invalid sort field",
                "allowed_values": VALID_APPROVAL_SORT_FIELDS
            }
        )

    if sort_order not in ["asc", "desc"]:
        raise HTTPException(
            status_code=422,
            detail="sort_order must be asc or desc"
        )

    query = db.query(Approval)

    if status:
        query = query.filter(
            Approval.status == status
        )

    if reviewer_id:
        query = query.filter(
            Approval.reviewer_id == reviewer_id
        )

    if decision_id:
        query = query.filter(
            Approval.decision_id == decision_id
        )

    if approval_level:
        query = query.filter(
            Approval.approval_level == approval_level
        )

    if start_date:
        query = query.filter(
            Approval.assigned_at >=
            datetime.combine(
                start_date,
                datetime.min.time()
            )
        )

    end_datetime = date_end_datetime(
        end_date
    )

    if end_datetime:
        query = query.filter(
            Approval.assigned_at < end_datetime
        )

    if sort_by == "approval_date":
        sort_column = Approval.completed_at
    else:
        sort_column = getattr(
            Approval,
            sort_by
        )

    if sort_order == "asc":
        query = query.order_by(
            sort_column.asc()
        )
    else:
        query = query.order_by(
            sort_column.desc()
        )

    total_records = query.count()

    approvals = query.offset(
        (page - 1) * page_size
    ).limit(
        page_size
    ).all()

    all_approvals = query.order_by(None).all()

    pending_count = sum(
        1 for a in all_approvals
        if a.status and a.status.lower() == "pending"
    )

    approved_count = sum(
        1 for a in all_approvals
        if a.status and a.status.lower() == "approved"
    )

    rejected_count = sum(
        1 for a in all_approvals
        if a.status and a.status.lower() == "rejected"
    )

    completed_approvals = [
        a for a in all_approvals
        if a.completed_at is not None
    ]

    turnaround_values = []

    for approval in completed_approvals:
        if approval.assigned_at and approval.completed_at:
            seconds = (
                approval.completed_at
                - approval.assigned_at
            ).total_seconds()

            turnaround_values.append(seconds)

    average_turnaround = (
        sum(turnaround_values) / len(turnaround_values)
        if turnaround_values
        else 0
    )

    completion_rate = (
        (
            len(completed_approvals)
            / len(all_approvals)
        ) * 100
        if all_approvals
        else 0
    )

    rows = []

    for approval in approvals:

        turnaround = None

        if approval.assigned_at and approval.completed_at:
            turnaround = int(
                (
                    approval.completed_at
                    - approval.assigned_at
                ).total_seconds()
            )

        rows.append({
            "approval_id": approval.id,
            "decision_id": approval.decision_id,
            "title": (
                approval.decision.title
                if approval.decision
                else None
            ),
            "reviewer": (
                approval.reviewer.full_name
                if approval.reviewer
                else None
            ),
            "reviewer_id": approval.reviewer_id,
            "approval_level": approval.approval_level,
            "status": approval.status,
            "assigned_date": approval.assigned_at,
            "completed_date": approval.completed_at,
            "turnaround_seconds": turnaround
        })

    return {
        "page": page,
        "page_size": page_size,
        "total_records": total_records,
        "summary": {
            "total": len(all_approvals),
            "pending": pending_count,
            "approved": approved_count,
            "rejected": rejected_count,
            "average_turnaround_seconds": round(
                average_turnaround,
                2
            ),
            "completion_rate": round(
                completion_rate,
                2
            )
        },
        "data": rows
    }
# ============================================================
# TEAM REPORT
# ============================================================

@router.get("/teams")
def get_team_report(
    team: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    sort_by: str = Query("team_name"),
    sort_order: str = Query("asc"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    if sort_by != "team_name":
        raise HTTPException(
            status_code=422,
            detail="Invalid sort field"
        )

    if sort_order.lower() not in ["asc", "desc"]:
        raise HTTPException(
            status_code=422,
            detail="Invalid sort order"
        )

    query = db.query(User)

    if team:
        query = query.filter(
            User.department == team
        )

    users = query.all()

    teams = {}

    for user in users:

        team_name = user.department or "Unassigned"

        if team_name not in teams:
            teams[team_name] = {
                "members": 0,
                "decisions": 0,
                "approved": 0,
                "rejected": 0,
                "pending": 0
            }

        teams[team_name]["members"] += 1

        decisions = (
            db.query(Decision)
            .filter(
                Decision.created_by == user.id
            )
        )

        if start_date:
            decisions = decisions.filter(
                Decision.created_at >= start_date
            )

        if end_date:
            decisions = decisions.filter(
                Decision.created_at < datetime.combine(
                    end_date,
                    datetime.max.time()
                )
            )

        if status:
            decisions = decisions.filter(
                Decision.status == status
            )

        if category:
            decisions = decisions.filter(
                Decision.category == category
            )

        decision_list = decisions.all()

        teams[team_name]["decisions"] += len(
            decision_list
        )

        for decision in decision_list:

            if decision.status == "Approved":
                teams[team_name]["approved"] += 1

            elif decision.status == "Rejected":
                teams[team_name]["rejected"] += 1

            else:
                teams[team_name]["pending"] += 1

    team_rows = []

    for team_name, values in teams.items():

        approval_total = (
            values["approved"]
            + values["rejected"]
        )

        approval_rate = (
            values["approved"]
            / approval_total
            * 100
            if approval_total
            else 0
        )

        team_rows.append({
            "team_name": team_name,
            "members": values["members"],
            "decisions": values["decisions"],
            "approved": values["approved"],
            "rejected": values["rejected"],
            "pending": values["pending"],
            "approval_total": approval_total,
            "approval_approved": values["approved"],
            "approval_rejected": values["rejected"],
            "approval_pending": values["pending"],
            "approval_rate": round(
                approval_rate,
                2
            )
        })

    reverse = sort_order.lower() == "desc"

    team_rows.sort(
        key=lambda x: x["team_name"],
        reverse=reverse
    )

    total = len(team_rows)

    start = (page - 1) * page_size
    end = start + page_size

    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "data": team_rows[start:end]
    }
    # ============================================================
# AUDIT REPORT
# ============================================================

@router.get("/audit")
def get_audit_report(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    # Only Manager and Administrator can access audit reports
    if current_user.role not in [
        "Manager",
        "Administrator"
    ]:
        raise HTTPException(
            status_code=403,
            detail="You do not have permission to access audit reports"
        )

    query = db.query(AuditLog)

    if user_id:
        query = query.filter(
            AuditLog.user_id == user_id
        )

    if action:
        query = query.filter(
            AuditLog.action == action
        )

    if entity_type:
        query = query.filter(
            AuditLog.entity_type == entity_type
        )

    if entity_id:
        query = query.filter(
            AuditLog.entity_id == entity_id
        )

    if start_date:
        query = query.filter(
            AuditLog.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            AuditLog.created_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    total = query.count()

    audit_logs = (
        query
        .order_by(AuditLog.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    rows = []

    for log in audit_logs:
        rows.append({
            "user": (
                log.user.full_name
                if log.user
                else None
            ),
            "user_id": log.user_id,
            "action": log.action,
            "entity_type": log.entity_type,
            "entity_id": log.entity_id,
            "description": log.description,
            "timestamp": log.created_at,
            "ip_address": log.ip_address
        })

    return {
        "page": page,
        "page_size": page_size,
        "total": total,
        "data": rows
    }
    # ============================================================
# DECISION REPORT - PDF EXPORT
# ============================================================

@router.get("/decisions/export/pdf")
def export_decisions_pdf(
    category: Optional[str] = None,
    status: Optional[str] = None,
    created_by: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    tag: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(Decision)

    if category:
        query = query.filter(Decision.category == category)

    if status:
        query = query.filter(Decision.status == status)

    if created_by:
        query = query.filter(
            Decision.created_by == created_by
        )

    if start_date:
        query = query.filter(
            Decision.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            Decision.created_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    decisions = query.order_by(
        Decision.created_at.desc()
    ).all()

    headers = [
        "ID",
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date"
    ]

    rows = []

    for decision in decisions:
        rows.append([
            decision.id,
            decision.title,
            decision.category,
            decision.status,
            decision.created_by,
            decision.created_at
        ])

    summary = {
        "Total": len(decisions),
        "Draft": sum(
            1 for d in decisions
            if d.status == "Draft"
        ),
        "Under Review": sum(
            1 for d in decisions
            if d.status == "Under Review"
        ),
        "Approved": sum(
            1 for d in decisions
            if d.status == "Approved"
        ),
        "Rejected": sum(
            1 for d in decisions
            if d.status == "Rejected"
        ),
        "Archived": sum(
            1 for d in decisions
            if d.status == "Archived"
        )
    }

    filters = {
        "Category": category,
        "Status": status,
        "Created By": created_by,
        "Start Date": start_date,
        "End Date": end_date,
        "Tag": tag
    }

    pdf_file = create_pdf(
        "Decision Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=decision_report.pdf"
        }
    )


# ============================================================
# APPROVAL REPORT - PDF EXPORT
# ============================================================

@router.get("/approvals/export/pdf")
def export_approvals_pdf(
    status: Optional[str] = None,
    reviewer_id: Optional[int] = None,
    decision_id: Optional[int] = None,
    approval_level: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(Approval)

    if status:
        query = query.filter(
            Approval.status == status
        )

    if reviewer_id:
        query = query.filter(
            Approval.reviewer_id == reviewer_id
        )

    if decision_id:
        query = query.filter(
            Approval.decision_id == decision_id
        )

    if approval_level:
        query = query.filter(
            Approval.approval_level == approval_level
        )

    if start_date:
        query = query.filter(
            Approval.assigned_at >= start_date
        )

    if end_date:
        query = query.filter(
            Approval.assigned_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    approvals = query.order_by(
        Approval.assigned_at.desc()
    ).all()

    headers = [
        "Approval ID",
        "Decision ID",
        "Title",
        "Reviewer",
        "Level",
        "Status",
        "Assigned Date",
        "Completed Date"
    ]

    rows = []

    for approval in approvals:

        rows.append([
            approval.id,
            approval.decision_id,
            approval.decision.title
            if approval.decision else "",
            approval.reviewer.full_name
            if approval.reviewer else "",
            approval.approval_level,
            approval.status,
            approval.assigned_at,
            approval.completed_at
        ])

    completed = [
        a for a in approvals
        if a.completed_at is not None
    ]

    pending = sum(
        1 for a in approvals
        if a.status == "Pending"
    )

    approved = sum(
        1 for a in approvals
        if a.status == "Approved"
    )

    rejected = sum(
        1 for a in approvals
        if a.status == "Rejected"
    )

    turnaround = []

    for approval in completed:
        if approval.assigned_at:
            turnaround.append(
                (
                    approval.completed_at
                    - approval.assigned_at
                ).total_seconds()
            )

    average_turnaround = (
        sum(turnaround) / len(turnaround)
        if turnaround
        else 0
    )

    completion_rate = (
        len(completed) / len(approvals) * 100
        if approvals
        else 0
    )

    summary = {
        "Total": len(approvals),
        "Pending": pending,
        "Approved": approved,
        "Rejected": rejected,
        "Average Turnaround Seconds":
            round(average_turnaround, 2),
        "Completion Rate":
            round(completion_rate, 2)
    }

    filters = {
        "Status": status,
        "Reviewer": reviewer_id,
        "Decision": decision_id,
        "Approval Level": approval_level,
        "Start Date": start_date,
        "End Date": end_date
    }

    pdf_file = create_pdf(
        "Approval Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=approval_report.pdf"
        }
    )
    # ============================================================
# TEAM REPORT - PDF EXPORT
# ============================================================

@router.get("/teams/export/pdf")
def export_teams_pdf(
    team: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(User)

    if team:
        query = query.filter(
            User.department == team
        )

    users = query.all()

    team_data = {}

    for user in users:

        team_name = user.department or "Unassigned"

        if team_name not in team_data:
            team_data[team_name] = {
                "members": 0,
                "decisions": 0,
                "approved": 0,
                "rejected": 0,
                "pending": 0
            }

        team_data[team_name]["members"] += 1

        decision_query = (
            db.query(Decision)
            .filter(
                Decision.created_by == user.id
            )
        )

        if start_date:
            decision_query = decision_query.filter(
                Decision.created_at >= start_date
            )

        if end_date:
            decision_query = decision_query.filter(
                Decision.created_at < datetime.combine(
                    end_date,
                    datetime.max.time()
                )
            )

        if status:
            decision_query = decision_query.filter(
                Decision.status == status
            )

        if category:
            decision_query = decision_query.filter(
                Decision.category == category
            )

        decisions = decision_query.all()

        team_data[team_name]["decisions"] += len(
            decisions
        )

        for decision in decisions:

            if decision.status == "Approved":
                team_data[team_name]["approved"] += 1

            elif decision.status == "Rejected":
                team_data[team_name]["rejected"] += 1

            else:
                team_data[team_name]["pending"] += 1

    headers = [
        "Team",
        "Members",
        "Decisions",
        "Approved",
        "Rejected",
        "Pending",
        "Approval Rate"
    ]

    rows = []

    for team_name, values in team_data.items():

        completed = (
            values["approved"]
            + values["rejected"]
        )

        approval_rate = (
            values["approved"]
            / completed
            * 100
            if completed
            else 0
        )

        rows.append([
            team_name,
            values["members"],
            values["decisions"],
            values["approved"],
            values["rejected"],
            values["pending"],
            round(approval_rate, 2)
        ])

    filters = {
        "Team": team,
        "Start Date": start_date,
        "End Date": end_date,
        "Status": status,
        "Category": category
    }

    pdf_file = create_pdf(
        "Team Report",
        headers,
        rows,
        filters=filters
    )

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=team_report.pdf"
        }
    )


# ============================================================
# AUDIT REPORT - PDF EXPORT
# ============================================================

@router.get("/audit/export/pdf")
def export_audit_pdf(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    if current_user.role not in [
        "Manager",
        "Administrator"
    ]:
        raise HTTPException(
            status_code=403,
            detail="You do not have permission to export audit reports"
        )

    query = db.query(AuditLog)

    if user_id:
        query = query.filter(
            AuditLog.user_id == user_id
        )

    if action:
        query = query.filter(
            AuditLog.action == action
        )

    if entity_type:
        query = query.filter(
            AuditLog.entity_type == entity_type
        )

    if entity_id:
        query = query.filter(
            AuditLog.entity_id == entity_id
        )

    if start_date:
        query = query.filter(
            AuditLog.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            AuditLog.created_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    audit_logs = (
        query
        .order_by(AuditLog.created_at.desc())
        .all()
    )

    headers = [
        "User",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "Timestamp",
        "IP Address"
    ]

    rows = []

    for log in audit_logs:

        rows.append([
            log.user.full_name
            if log.user else "",
            log.action,
            log.entity_type,
            log.entity_id,
            log.description,
            log.created_at,
            log.ip_address
        ])

    summary = {
        "Total Audit Records": len(audit_logs)
    }

    filters = {
        "User": user_id,
        "Action": action,
        "Entity Type": entity_type,
        "Entity ID": entity_id,
        "Start Date": start_date,
        "End Date": end_date
    }

    pdf_file = create_pdf(
        "Audit Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        pdf_file,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=audit_report.pdf"
        }
    )
    # ============================================================
# DECISION REPORT - EXCEL EXPORT
# ============================================================

@router.get("/decisions/export/excel")
def export_decisions_excel(
    category: Optional[str] = None,
    status: Optional[str] = None,
    created_by: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    tag: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(Decision)

    if category:
        query = query.filter(Decision.category == category)

    if status:
        query = query.filter(Decision.status == status)

    if created_by:
        query = query.filter(
            Decision.created_by == created_by
        )

    if start_date:
        query = query.filter(
            Decision.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            Decision.created_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    decisions = query.order_by(
        Decision.created_at.desc()
    ).all()

    headers = [
        "Decision ID",
        "Title",
        "Category",
        "Status",
        "Created By",
        "Created Date",
        "Updated Date"
    ]

    rows = []

    for decision in decisions:
        rows.append([
            decision.id,
            decision.title,
            decision.category,
            decision.status,
            decision.created_by,
            decision.created_at,
            decision.updated_at
        ])

    summary = {
        "Total": len(decisions),
        "Draft": sum(
            1 for d in decisions
            if d.status == "Draft"
        ),
        "Under Review": sum(
            1 for d in decisions
            if d.status == "Under Review"
        ),
        "Approved": sum(
            1 for d in decisions
            if d.status == "Approved"
        ),
        "Rejected": sum(
            1 for d in decisions
            if d.status == "Rejected"
        ),
        "Archived": sum(
            1 for d in decisions
            if d.status == "Archived"
        )
    }

    filters = {
        "Category": category,
        "Status": status,
        "Created By": created_by,
        "Start Date": start_date,
        "End Date": end_date,
        "Tag": tag
    }

    excel_file = create_excel(
        "Decision Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        excel_file,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=decision_report.xlsx"
        }
    )


# ============================================================
# APPROVAL REPORT - EXCEL EXPORT
# ============================================================

@router.get("/approvals/export/excel")
def export_approvals_excel(
    status: Optional[str] = None,
    reviewer_id: Optional[int] = None,
    decision_id: Optional[int] = None,
    approval_level: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(Approval)

    if status:
        query = query.filter(
            Approval.status == status
        )

    if reviewer_id:
        query = query.filter(
            Approval.reviewer_id == reviewer_id
        )

    if decision_id:
        query = query.filter(
            Approval.decision_id == decision_id
        )

    if approval_level:
        query = query.filter(
            Approval.approval_level == approval_level
        )

    if start_date:
        query = query.filter(
            Approval.assigned_at >= start_date
        )

    if end_date:
        query = query.filter(
            Approval.assigned_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    approvals = query.order_by(
        Approval.assigned_at.desc()
    ).all()

    headers = [
        "Approval ID",
        "Decision ID",
        "Title",
        "Reviewer",
        "Reviewer ID",
        "Approval Level",
        "Status",
        "Assigned Date",
        "Completed Date",
        "Turnaround Seconds"
    ]

    rows = []

    for approval in approvals:

        turnaround = None

        if (
            approval.assigned_at
            and approval.completed_at
        ):
            turnaround = int(
                (
                    approval.completed_at
                    - approval.assigned_at
                ).total_seconds()
            )

        rows.append([
            approval.id,
            approval.decision_id,
            approval.decision.title
            if approval.decision else "",
            approval.reviewer.full_name
            if approval.reviewer else "",
            approval.reviewer_id,
            approval.approval_level,
            approval.status,
            approval.assigned_at,
            approval.completed_at,
            turnaround
        ])

    completed = [
        a for a in approvals
        if a.completed_at is not None
    ]

    summary = {
        "Total": len(approvals),
        "Pending": sum(
            1 for a in approvals
            if a.status == "Pending"
        ),
        "Approved": sum(
            1 for a in approvals
            if a.status == "Approved"
        ),
        "Rejected": sum(
            1 for a in approvals
            if a.status == "Rejected"
        ),
        "Completed": len(completed)
    }

    filters = {
        "Status": status,
        "Reviewer": reviewer_id,
        "Decision": decision_id,
        "Approval Level": approval_level,
        "Start Date": start_date,
        "End Date": end_date
    }

    excel_file = create_excel(
        "Approval Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        excel_file,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=approval_report.xlsx"
        }
    )
    # ============================================================
# TEAM REPORT - EXCEL EXPORT
# ============================================================

@router.get("/teams/export/excel")
def export_teams_excel(
    team: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    category: Optional[str] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    query = db.query(User)

    if team:
        query = query.filter(
            User.department == team
        )

    users = query.all()

    team_data = {}

    for user in users:

        team_name = user.department or "Unassigned"

        if team_name not in team_data:
            team_data[team_name] = {
                "members": 0,
                "decisions": 0,
                "approved": 0,
                "rejected": 0,
                "pending": 0
            }

        team_data[team_name]["members"] += 1

        decision_query = (
            db.query(Decision)
            .filter(
                Decision.created_by == user.id
            )
        )

        if start_date:
            decision_query = decision_query.filter(
                Decision.created_at >= start_date
            )

        if end_date:
            decision_query = decision_query.filter(
                Decision.created_at < datetime.combine(
                    end_date,
                    datetime.max.time()
                )
            )

        if status:
            decision_query = decision_query.filter(
                Decision.status == status
            )

        if category:
            decision_query = decision_query.filter(
                Decision.category == category
            )

        decisions = decision_query.all()

        team_data[team_name]["decisions"] += len(
            decisions
        )

        for decision in decisions:

            if decision.status == "Approved":
                team_data[team_name]["approved"] += 1

            elif decision.status == "Rejected":
                team_data[team_name]["rejected"] += 1

            else:
                team_data[team_name]["pending"] += 1

    headers = [
        "Team",
        "Members",
        "Decisions",
        "Approved",
        "Rejected",
        "Pending",
        "Approval Rate"
    ]

    rows = []

    for team_name, values in team_data.items():

        completed = (
            values["approved"]
            + values["rejected"]
        )

        approval_rate = (
            values["approved"]
            / completed
            * 100
            if completed
            else 0
        )

        rows.append([
            team_name,
            values["members"],
            values["decisions"],
            values["approved"],
            values["rejected"],
            values["pending"],
            round(approval_rate, 2)
        ])

    filters = {
        "Team": team,
        "Start Date": start_date,
        "End Date": end_date,
        "Status": status,
        "Category": category
    }

    excel_file = create_excel(
        "Team Report",
        headers,
        rows,
        filters=filters
    )

    return StreamingResponse(
        excel_file,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=team_report.xlsx"
        }
    )


# ============================================================
# AUDIT REPORT - EXCEL EXPORT
# ============================================================

@router.get("/audit/export/excel")
def export_audit_excel(
    user_id: Optional[int] = None,
    action: Optional[str] = None,
    entity_type: Optional[str] = None,
    entity_id: Optional[int] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):

    if current_user.role not in [
        "Manager",
        "Administrator"
    ]:
        raise HTTPException(
            status_code=403,
            detail="You do not have permission to export audit reports"
        )

    query = db.query(AuditLog)

    if user_id:
        query = query.filter(
            AuditLog.user_id == user_id
        )

    if action:
        query = query.filter(
            AuditLog.action == action
        )

    if entity_type:
        query = query.filter(
            AuditLog.entity_type == entity_type
        )

    if entity_id:
        query = query.filter(
            AuditLog.entity_id == entity_id
        )

    if start_date:
        query = query.filter(
            AuditLog.created_at >= start_date
        )

    if end_date:
        query = query.filter(
            AuditLog.created_at < datetime.combine(
                end_date,
                datetime.max.time()
            )
        )

    audit_logs = (
        query
        .order_by(AuditLog.created_at.desc())
        .all()
    )

    headers = [
        "User",
        "User ID",
        "Action",
        "Entity Type",
        "Entity ID",
        "Description",
        "Timestamp",
        "IP Address"
    ]

    rows = []

    for log in audit_logs:

        rows.append([
            log.user.full_name
            if log.user else "",
            log.user_id,
            log.action,
            log.entity_type,
            log.entity_id,
            log.description,
            log.created_at,
            log.ip_address
        ])

    summary = {
        "Total Audit Records": len(audit_logs)
    }

    filters = {
        "User": user_id,
        "Action": action,
        "Entity Type": entity_type,
        "Entity ID": entity_id,
        "Start Date": start_date,
        "End Date": end_date
    }

    excel_file = create_excel(
        "Audit Report",
        headers,
        rows,
        summary,
        filters
    )

    return StreamingResponse(
        excel_file,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            "attachment; filename=audit_report.xlsx"
        }
    )
